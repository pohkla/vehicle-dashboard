from __future__ import annotations

import os
import re
import unicodedata
import json
import base64
import time
import urllib.request
import urllib.error
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse

ADMIN_TOKEN = os.getenv("ADMIN_TOKEN", "change-this-token")
GITHUB_TOKEN = os.getenv("GITHUB_TOKEN", "")
GITHUB_REPO = os.getenv("GITHUB_REPO", "")
GITHUB_FILE = os.getenv("GITHUB_FILE", "data.json")
GITHUB_BRANCH = os.getenv("GITHUB_BRANCH", "main")
CACHE_TTL_SECONDS = int(os.getenv("CACHE_TTL_SECONDS", "20"))
DASHBOARD_CACHE: dict[str, Any] = {"key": None, "data": None, "created_at": 0.0}

APP_VERSION = "v14.21 UIUX Professional Polish"
app = FastAPI(title=f"Vehicle Dashboard {APP_VERSION}")


def github_enabled() -> bool:
    return bool(GITHUB_TOKEN and GITHUB_REPO and GITHUB_FILE)


def github_api_url() -> str:
    return f"https://api.github.com/repos/{GITHUB_REPO}/contents/{GITHUB_FILE}"


def github_headers() -> dict[str, str]:
    return {
        "Authorization": f"Bearer {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
        "User-Agent": "vehicle-dashboard",
    }


def github_request(method: str, url: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    data = None
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(url, data=data, headers=github_headers(), method=method)
    try:
        with urllib.request.urlopen(req, timeout=30) as res:
            body = res.read().decode("utf-8")
            return json.loads(body) if body else {}
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore")
        raise HTTPException(status_code=500, detail=f"GitHub API error {e.code}: {detail}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"GitHub connection error: {str(e)}")


def empty_store() -> dict[str, Any]:
    return {"version": 1, "updated_at": None, "daily_records": [], "weekly_summaries": []}


def read_github_store() -> tuple[dict[str, Any], str | None]:
    if not github_enabled():
        raise HTTPException(status_code=500, detail="ยังไม่ได้ตั้งค่า GITHUB_TOKEN / GITHUB_REPO / GITHUB_FILE")
    url = github_api_url() + f"?ref={GITHUB_BRANCH}"
    try:
        res = github_request("GET", url)
    except HTTPException as e:
        if "GitHub API error 404" in str(e.detail):
            return empty_store(), None
        raise
    content = base64.b64decode(res.get("content", "")).decode("utf-8")
    data = json.loads(content) if content.strip() else empty_store()
    data.setdefault("daily_records", [])
    data.setdefault("weekly_summaries", [])
    return data, res.get("sha")


def write_github_store(data: dict[str, Any], sha: str | None, message: str) -> None:
    content = base64.b64encode(json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")).decode("utf-8")
    payload: dict[str, Any] = {"message": message, "content": content, "branch": GITHUB_BRANCH}
    if sha:
        payload["sha"] = sha
    github_request("PUT", github_api_url(), payload)


def make_cache_key(start: str | None, end: str | None, q: str | None) -> str:
    return f"start={start or ''}|end={end or ''}|q={q or ''}"


def clear_dashboard_cache() -> None:
    DASHBOARD_CACHE["key"] = None
    DASHBOARD_CACHE["data"] = None
    DASHBOARD_CACHE["created_at"] = 0.0


def get_cached_dashboard(key: str) -> dict[str, Any] | None:
    if DASHBOARD_CACHE["key"] != key or DASHBOARD_CACHE["data"] is None:
        return None
    if time.time() - float(DASHBOARD_CACHE["created_at"] or 0) > CACHE_TTL_SECONDS:
        return None
    return DASHBOARD_CACHE["data"]


def set_cached_dashboard(key: str, data: dict[str, Any]) -> None:
    DASHBOARD_CACHE["key"] = key
    DASHBOARD_CACHE["data"] = data
    DASHBOARD_CACHE["created_at"] = time.time()


@app.on_event("startup")
def startup() -> None:
    pass


def only_number(value: str) -> float:
    allowed = "0123456789."
    text = "".join(ch for ch in value.replace(",", "") if ch in allowed)
    return float(text) if text else 0


def extract_date(line: str) -> str:
    for part in line.replace("วันที่", " ").replace("ที่", " ").split():
        if "/" in part:
            bits = part.split("/")
            if len(bits) == 3:
                day, month, year = bits
                if day.isdigit() and month.isdigit() and year.isdigit():
                    return f"{int(day):02d}/{int(month):02d}/{year}"
    return "ไม่พบวันที่"


def thai_date_to_iso(date_text: str) -> str:
    if "/" not in date_text:
        return ""
    day, month, year = date_text.split("/")
    y = int(year)
    if y > 2400:
        y -= 543
    return f"{y:04d}-{int(month):02d}-{int(day):02d}"


def vehicle_meta(line: str) -> dict[str, str] | None:
    if "รถจักรยานยนต์" in line:
        return {"key": "motorcycle", "icon": "🏍", "title": "รถจักรยานยนต์"}
    if "รถกระบะ" in line:
        return {"key": "pickup", "icon": "🚛", "title": "รถกระบะ"}
    if "รถยนต์เก๋ง" in line:
        return {"key": "sedan", "icon": "🚗", "title": "รถยนต์เก๋ง"}
    return None


def parse_period_key(line: str) -> str:
    return line.replace("#", "").strip()


def parse_report(raw_text: str) -> dict[str, Any]:
    lines = [line.strip() for line in raw_text.replace(chr(13), "").splitlines()]
    rows: list[dict[str, Any]] = []
    weekly_summaries: dict[str, dict[str, int]] = {}

    current_date_text = ""
    current_iso_date = ""
    current_meta: dict[str, str] | None = None
    current_company = ""
    current_period = ""

    for line in lines:
        if not line or line.startswith("====") or line.startswith("####"):
            continue

        if "รายสัปดาห์" in line and any(ch.isdigit() for ch in line):
            current_period = parse_period_key(line)
            weekly_summaries.setdefault(
                current_period,
                {"car": 0, "motorcycle": 0, "total": 0},
            )
            continue

        if current_period and "บาท" in line:
            target = line.split("->", 1)[1] if "->" in line else line
            value = round(only_number(target))
            if value:
                if "รวม" in line:
                    weekly_summaries[current_period]["total"] = value
                elif "รถจักรยานยนต์" in line:
                    weekly_summaries[current_period]["motorcycle"] = value
                elif "รถยนต์" in line:
                    weekly_summaries[current_period]["car"] = value
            continue

        if "สรุปยอด" in line and "/" in line:
            current_date_text = extract_date(line)
            current_iso_date = thai_date_to_iso(current_date_text)
            current_meta = None
            current_company = ""
            continue

        if not current_iso_date:
            continue

        meta = vehicle_meta(line)
        if meta:
            current_meta = meta
            current_company = ""
            continue

        if line.startswith("[") and line.endswith("]"):
            current_company = line[1:-1].strip()
            continue

        if line.startswith("•") and current_meta:
            item_text = line[1:].strip().replace("_", " ")
            rows.append(
                {
                    "date": current_date_text,
                    "isoDate": current_iso_date,
                    "vehicleType": current_meta["key"],
                    "vehicleTitle": current_meta["title"],
                    "icon": current_meta["icon"],
                    "company": current_company or "",
                    "item": item_text,
                }
            )

    for period, summary in weekly_summaries.items():
        if not summary["total"]:
            summary["total"] = summary["car"] + summary["motorcycle"]

    return {
        "rows": rows,
        "weeklySummaries": weekly_summaries,
    }


def save_import_replace_all(raw_text: str) -> dict[str, int]:
    parsed = parse_report(raw_text)
    rows = parsed["rows"]
    weekly_summaries = parsed["weeklySummaries"]
    if not rows:
        raise HTTPException(status_code=400, detail="อ่านข้อมูลไม่สำเร็จ: ไม่พบรายการรายวัน")

    now = datetime.utcnow().isoformat()
    imported_dates = sorted({row["isoDate"] for row in rows if row["isoDate"]})
    store, sha = read_github_store()
    old_record_count = len(store.get("daily_records", []))
    old_summary_count = len(store.get("weekly_summaries", []))

    records = []
    for row in rows:
        records.append({
            "date_text": row["date"],
            "iso_date": row["isoDate"],
            "vehicle_type": row["vehicleType"],
            "vehicle_title": row["vehicleTitle"],
            "icon": row["icon"],
            "company": row["company"],
            "item": row["item"],
            "created_at": now,
        })

    summaries = []
    for period, summary in weekly_summaries.items():
        summaries.append({
            "period_key": period,
            "car_amount": summary["car"],
            "motorcycle_amount": summary["motorcycle"],
            "total_amount": summary["total"],
            "updated_at": now,
        })

    new_store = {"version": 1, "updated_at": now, "daily_records": records, "weekly_summaries": summaries}
    write_github_store(new_store, sha, f"update vehicle dashboard data {now}")
    clear_dashboard_cache()

    return {
        "report_id": 0,
        "imported_dates": len(imported_dates),
        "deleted_records": int(old_record_count or 0),
        "deleted_summaries": int(old_summary_count or 0),
        "inserted": len(records),
        "replaced_summaries": len(summaries),
        "parsed_rows": len(rows),
        "duplicated": 0,
    }



# -----------------------------
# Excel Import (stable add-on for v6.2)
# -----------------------------
def normalize_vehicle_type(value: Any) -> dict[str, str]:
    text = str(value or "").strip().replace(" ", "")
    if "มอเตอร์" in text or "จักรยานยนต์" in text:
        return {"key": "motorcycle", "icon": "🏍", "title": "รถจักรยานยนต์"}
    if "กระบะ" in text:
        return {"key": "pickup", "icon": "🚛", "title": "รถกระบะ"}
    if "เก๋ง" in text or "รถยนต์" in text:
        return {"key": "sedan", "icon": "🚗", "title": "รถยนต์เก๋ง"}
    return {"key": "unknown", "icon": "", "title": str(value or "").strip() or "ไม่ระบุ"}


def normalize_company(value: Any, vehicle_key: str = "") -> str:
    text = str(value or "").strip()
    low = text.lower()
    if vehicle_key == "motorcycle" and not text:
        return "บริษัทกลางฯ RVP"
    if "rvp" in low or "บริษัทกลาง" in text:
        return "บริษัทกลางฯ RVP"
    if "ergo" in low:
        return "ERGO"
    if "ไทยไพบูลย์" in text or "tpb" in low:
        return "ไทยไพบูลย์ TPB"
    return text


def excel_date_to_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if hasattr(value, "day") and hasattr(value, "month") and hasattr(value, "year"):
        return f"{int(value.day):02d}/{int(value.month):02d}/{int(value.year)}"
    text = str(value).strip()
    if not text:
        return ""
    if "/" in text:
        parts = text.split("/")
        if len(parts) == 3:
            return f"{int(float(parts[0])):02d}/{int(float(parts[1])):02d}/{int(float(parts[2]))}"
    return text


def to_money(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).replace(",", "").strip()
    keep = "".join(ch for ch in text if ch.isdigit() or ch in ".-")
    try:
        return round(float(keep), 2) if keep else 0.0
    except ValueError:
        return 0.0



def parse_expense_summaries_from_sheet(ws) -> list[dict[str, Any]]:
    """Read the expense summary area on the right side of the Excel sheet.

    Supported layouts seen in the user's files:
    - a heading cell containing "ยอดค่าใช้จ่าย"
    - rows below with labels like "รถมอเตอร์ไซค์" and "รถยนต์"
    - amount can be in งวด 1, งวด 2, or any numeric cell to the right of the label

    We intentionally keep this parser separate from the main A-G table parser,
    because the old import flow skipped the right-side summary area by design.
    """
    summaries: list[dict[str, Any]] = []
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    heading_cells: list[tuple[int, int]] = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            text = str(ws.cell(row, col).value or "").strip().replace(" ", "")
            if "ยอดค่าใช้จ่าย" in text:
                heading_cells.append((row, col))

    def row_amount(row_idx: int, start_col: int) -> float:
        total = 0.0
        # Expense tables are usually 3-5 columns wide, but keep a wider safe window.
        for col in range(start_col + 1, min(max_col, start_col + 8) + 1):
            val = ws.cell(row_idx, col).value
            # Ignore count columns and text/unit cells by using to_money; blank/text => 0.
            total += to_money(val)
        return round(total, 2)

    for head_row, head_col in heading_cells:
        motorcycle_amount = 0.0
        car_amount = 0.0
        # Scan rows below the heading and a few columns around the heading.
        for row in range(head_row + 1, min(max_row, head_row + 12) + 1):
            for col in range(max(1, head_col - 2), min(max_col, head_col + 5) + 1):
                label = str(ws.cell(row, col).value or "").strip().replace(" ", "")
                if not label:
                    continue
                amount = row_amount(row, col)
                if amount <= 0:
                    continue
                if "มอเตอร์" in label or "จักรยานยนต์" in label:
                    motorcycle_amount += amount
                    break
                if "รถยนต์" in label or "รายยนต์" in label:
                    car_amount += amount
                    break

        if motorcycle_amount or car_amount:
            summaries.append({
                "source_sheet": ws.title,
                "motorcycle_amount": round(motorcycle_amount, 2),
                "car_amount": round(car_amount, 2),
                "total_amount": round(motorcycle_amount + car_amount, 2),
            })

    return summaries

def parse_excel_upload(content: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
        from io import BytesIO
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ยังไม่ได้ติดตั้ง openpyxl: {e}")
    wb = load_workbook(BytesIO(content), data_only=True)
    required = ["วันที่", "ประเภทรถ", "บริษัท", "รหัส", "ยอดสุทธิ", "ยอดเก็บจริง"]
    records: list[dict[str, Any]] = []
    sheet_stats: list[dict[str, Any]] = []
    expense_summaries: list[dict[str, Any]] = []
    now = datetime.utcnow().isoformat()
    for ws in wb.worksheets:
        sheet_expenses = parse_expense_summaries_from_sheet(ws)
        expense_summaries.extend(sheet_expenses)
        found_header_row = None
        header_map: dict[str, int] = {}
        headers_seen: list[str] = []
        for row_idx in range(1, min(ws.max_row, 15) + 1):
            values = [ws.cell(row_idx, col).value for col in range(1, min(ws.max_column, 8) + 1)]
            normalized = [str(v).strip() if v is not None else "" for v in values]
            if "วันที่" in normalized and "ประเภทรถ" in normalized and "รหัส" in normalized:
                found_header_row = row_idx
                headers_seen = normalized
                for idx, name in enumerate(normalized, start=1):
                    if name in required and name not in header_map:
                        header_map[name] = idx
                break
        if not found_header_row:
            sheet_stats.append({"sheet": ws.title, "status": "skip", "reason": "ไม่พบ header รายการ", "rows": 0})
            continue
        missing = [h for h in required if h not in header_map]
        if missing:
            raise HTTPException(status_code=400, detail=f"ชีต {ws.title} ขาด header: {', '.join(missing)}")
        current_date_text = ""
        current_iso_date = ""
        start_count = len(records)
        for row_idx in range(found_header_row + 1, ws.max_row + 1):
            date_raw = ws.cell(row_idx, header_map["วันที่"]).value
            if date_raw not in (None, ""):
                current_date_text = excel_date_to_text(date_raw)
                current_iso_date = thai_date_to_iso(current_date_text)
            vehicle_raw = ws.cell(row_idx, header_map["ประเภทรถ"]).value
            code_raw = ws.cell(row_idx, header_map["รหัส"]).value
            if not current_iso_date or not vehicle_raw or not code_raw:
                continue
            meta = normalize_vehicle_type(vehicle_raw)
            if meta["key"] == "unknown":
                continue
            company = normalize_company(ws.cell(row_idx, header_map["บริษัท"]).value, meta["key"])
            records.append({
                "date_text": current_date_text,
                "iso_date": current_iso_date,
                "vehicle_type": meta["key"],
                "vehicle_title": meta["title"],
                "icon": meta["icon"],
                "company": company,
                "item": str(code_raw).strip(),
                "net_amount": to_money(ws.cell(row_idx, header_map["ยอดสุทธิ"]).value),
                "collected_amount": to_money(ws.cell(row_idx, header_map["ยอดเก็บจริง"]).value),
                "import_type": "excel",
                "source_sheet": ws.title,
                "created_at": now,
            })
        sheet_stats.append({"sheet": ws.title, "status": "parsed", "header_row": found_header_row, "headers": headers_seen, "rows": len(records) - start_count, "expense_summaries": sheet_expenses})
    if not records:
        raise HTTPException(status_code=400, detail="อ่าน Excel ไม่สำเร็จ: ไม่พบรายการข้อมูล")
    return {"records": records, "sheet_stats": sheet_stats, "parsed_rows": len(records), "expense_summaries": expense_summaries}


def save_excel_replace_all(file_content: bytes, filename: str = "") -> dict[str, Any]:
    parsed = parse_excel_upload(file_content)
    records = parsed["records"]
    now = datetime.utcnow().isoformat()
    imported_dates = sorted({row["iso_date"] for row in records if row.get("iso_date")})
    store, sha = read_github_store()
    old_record_count = len(store.get("daily_records", []))
    old_summary_count = len(store.get("weekly_summaries", []))
    car_net = sum(float(r.get("net_amount", 0) or 0) for r in records if r.get("vehicle_type") in ("pickup", "sedan"))
    motor_net = sum(float(r.get("net_amount", 0) or 0) for r in records if r.get("vehicle_type") == "motorcycle")
    weekly_summaries = [{"period_key": f"Excel Import {filename or now}", "car_amount": round(car_net, 2), "motorcycle_amount": round(motor_net, 2), "total_amount": round(car_net + motor_net, 2), "updated_at": now}]
    expense_summaries = parsed.get("expense_summaries", [])
    # Add each sheet date range for later dashboard filtering/allocation.
    for ex in expense_summaries:
        sheet_rows = [r for r in records if r.get("source_sheet") == ex.get("source_sheet") and r.get("iso_date")]
        dates = sorted({r.get("iso_date") for r in sheet_rows if r.get("iso_date")})
        ex["date_start"] = dates[0] if dates else None
        ex["date_end"] = dates[-1] if dates else None
        ex["updated_at"] = now
    new_store = {"version": 3, "app_version": APP_VERSION, "import_type": "excel", "updated_at": now, "daily_records": records, "weekly_summaries": weekly_summaries, "expense_summaries": expense_summaries, "excel_debug": {"filename": filename, "sheet_stats": parsed["sheet_stats"], "parsed_rows": parsed["parsed_rows"], "expense_summaries": expense_summaries, "required_money_fields_ok": all("net_amount" in r and "collected_amount" in r for r in records)}}
    write_github_store(new_store, sha, f"excel import vehicle dashboard data {now}")
    clear_dashboard_cache()
    verify_store, verify_sha = read_github_store()
    return {"report_id": 0, "import_type": "excel", "version": APP_VERSION, "imported_dates": len(imported_dates), "deleted_records": int(old_record_count or 0), "deleted_summaries": int(old_summary_count or 0), "inserted": len(records), "replaced_summaries": len(weekly_summaries), "parsed_rows": parsed["parsed_rows"], "sheet_stats": parsed["sheet_stats"], "github_write_verified": len(verify_store.get("daily_records", [])) == len(records), "sha_exists_after_write": bool(verify_sha)}


def get_money_totals_from_weekly_summaries(store: dict[str, Any]) -> dict[str, float]:
    summaries = store.get("weekly_summaries", [])
    car = sum(float(s.get("car_amount", 0) or 0) for s in summaries)
    motorcycle = sum(float(s.get("motorcycle_amount", 0) or 0) for s in summaries)
    total = sum(float(s.get("total_amount", 0) or 0) for s in summaries)
    # Old text-import data has only one amount field. Keep it as net for backward compatibility.
    return {
        "net": {"pickup": round(car, 2), "sedan": 0, "car": round(car, 2), "motorcycle": round(motorcycle, 2), "total": round(total, 2)},
        "collected": {"pickup": 0, "sedan": 0, "car": 0, "motorcycle": 0, "total": 0},
        "legacy": True,
    }


def get_expense_totals_from_records(rows: list[dict[str, Any]], store: dict[str, Any]) -> dict[str, float]:
    """Allocate Excel right-side expense summaries into the current dashboard filter.

    Expense summaries are sheet-level in the uploaded Excel. If the dashboard filter
    shows only part of a sheet, expenses are proportionally allocated by vehicle count.
    Motorcycle expense maps directly to motorcycle. Car expense is split into pickup
    and sedan by their counts in the same sheet.
    """
    expense = {"pickup": 0.0, "sedan": 0.0, "car": 0.0, "motorcycle": 0.0, "total": 0.0}
    expense_summaries = store.get("expense_summaries", []) or []
    if not expense_summaries:
        return {k: 0.0 for k in expense}

    all_rows = store.get("daily_records", []) or []
    filtered_by_sheet: dict[str, list[dict[str, Any]]] = {}
    all_by_sheet: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        filtered_by_sheet.setdefault(str(row.get("source_sheet") or ""), []).append(row)
    for row in all_rows:
        all_by_sheet.setdefault(str(row.get("source_sheet") or ""), []).append(row)

    def count_type(source: list[dict[str, Any]], types: tuple[str, ...]) -> int:
        return sum(1 for r in source if r.get("vehicle_type") in types)

    for summary in expense_summaries:
        sheet = str(summary.get("source_sheet") or "")
        if not sheet:
            continue
        sheet_all = all_by_sheet.get(sheet, [])
        sheet_filtered = filtered_by_sheet.get(sheet, [])
        if not sheet_all or not sheet_filtered:
            continue

        all_motor = count_type(sheet_all, ("motorcycle",))
        fil_motor = count_type(sheet_filtered, ("motorcycle",))
        all_pickup = count_type(sheet_all, ("pickup",))
        all_sedan = count_type(sheet_all, ("sedan",))
        fil_pickup = count_type(sheet_filtered, ("pickup",))
        fil_sedan = count_type(sheet_filtered, ("sedan",))
        all_car = all_pickup + all_sedan

        motor_amount = float(summary.get("motorcycle_amount", 0) or 0)
        car_amount = float(summary.get("car_amount", 0) or 0)

        if all_motor:
            expense["motorcycle"] += motor_amount * (fil_motor / all_motor)
        if all_car:
            expense["pickup"] += car_amount * (fil_pickup / all_car)
            expense["sedan"] += car_amount * (fil_sedan / all_car)

    expense["car"] = expense["pickup"] + expense["sedan"]
    expense["total"] = expense["car"] + expense["motorcycle"]
    return {k: round(v, 2) for k, v in expense.items()}


def get_money_totals_from_records(rows: list[dict[str, Any]], store: dict[str, Any], use_legacy_fallback: bool = False) -> dict[str, Any]:
    net = {"pickup": 0.0, "sedan": 0.0, "car": 0.0, "motorcycle": 0.0, "total": 0.0}
    collected = {"pickup": 0.0, "sedan": 0.0, "car": 0.0, "motorcycle": 0.0, "total": 0.0}
    has_money_fields = False

    for row in rows:
        vehicle_type = row.get("vehicle_type")
        bucket = "motorcycle" if vehicle_type == "motorcycle" else vehicle_type if vehicle_type in ("pickup", "sedan") else "car" if vehicle_type == "car" else None
        if bucket is None:
            continue

        if "net_amount" in row or "collected_amount" in row:
            has_money_fields = True
        net_amount = float(row.get("net_amount", 0) or 0)
        collected_amount = float(row.get("collected_amount", 0) or 0)

        net[bucket] += net_amount
        collected[bucket] += collected_amount
        if bucket in ("pickup", "sedan", "car"):
            net["car"] += net_amount
            collected["car"] += collected_amount
        net["total"] += net_amount
        collected["total"] += collected_amount

    if not has_money_fields and use_legacy_fallback:
        base = get_money_totals_from_weekly_summaries(store)
        base["expense"] = {"pickup": 0, "sedan": 0, "car": 0, "motorcycle": 0, "total": 0}
        base["remaining"] = base.get("collected", {}).copy()
        return base

    expense = get_expense_totals_from_records(rows, store)
    remaining = {k: round((collected.get(k, 0) or 0) - (expense.get(k, 0) or 0), 2) for k in collected}

    return {
        "net": {k: round(v, 2) for k, v in net.items()},
        "collected": {k: round(v, 2) for k, v in collected.items()},
        "expense": expense,
        "remaining": remaining,
        "legacy": False,
    }



def normalize_search_text(value: Any) -> str:
    """Normalize search text so vehicle plate/code variants match.

    Examples:
    - กษบ164 == กษบ-164
    - ผจ-4573 == ผจ4573
    - ผจ-4573สงขลา == ผจ-4573 สงขลา

    The function removes every character that is not a Unicode letter or number,
    so separators such as spaces, hyphens, dots, slashes, underscores,
    non-breaking spaces, and punctuation do not affect search matching.
    """
    text = unicodedata.normalize("NFKC", str(value or "")).strip().lower()
    return re.sub(r"[^\w]+", "", text, flags=re.UNICODE)

def get_dashboard_data(start: str | None = None, end: str | None = None, q: str | None = None) -> dict[str, Any]:
    store, _ = read_github_store()
    all_rows = store.get("daily_records", [])
    q_lower = (q or "").strip().lower()
    q_normalized = normalize_search_text(q_lower)

    filtered_rows = []
    for row in all_rows:
        iso_date = row.get("iso_date", "")
        if start and iso_date < start:
            continue
        if end and iso_date > end:
            continue
        if q_lower:
            haystack = " ".join([row.get("item", ""), row.get("company", ""), row.get("vehicle_title", ""), row.get("date_text", "")]).lower()
            haystack_normalized = normalize_search_text(haystack)
            if q_lower not in haystack and q_normalized not in haystack_normalized:
                continue
        filtered_rows.append(row)

    days: dict[str, Any] = {}
    for row in sorted(filtered_rows, key=lambda r: (r.get("iso_date", ""), r.get("item", ""))):
        day_key = row.get("iso_date", "")
        if day_key not in days:
            days[day_key] = {"date": row.get("date_text", ""), "isoDate": row.get("iso_date", ""), "motorcycle": 0, "pickup": 0, "sedan": 0, "groups": {}}
        day = days[day_key]
        vehicle_type = row.get("vehicle_type", "")
        company = row.get("company", "") or ""
        group_key = f"{vehicle_type}|{company}"
        if group_key not in day["groups"]:
            day["groups"][group_key] = {"key": vehicle_type, "icon": row.get("icon", ""), "title": row.get("vehicle_title", ""), "company": company, "items": []}
        day["groups"][group_key]["items"].append({
            "text": row.get("item", ""),
            "net_amount": float(row.get("net_amount", 0) or 0),
            "collected_amount": float(row.get("collected_amount", 0) or 0),
            "vehicle_type": vehicle_type,
            "vehicle_title": row.get("vehicle_title", ""),
            "company": company,
        })
        if vehicle_type in ("motorcycle", "pickup", "sedan"):
            day[vehicle_type] += 1

    daily_data = []
    for day in days.values():
        groups = []
        for group in day["groups"].values():
            group["count"] = len(group["items"])
            groups.append(group)
        day["groups"] = groups
        daily_data.append(day)

    motorcycle = sum(1 for r in filtered_rows if r.get("vehicle_type") == "motorcycle")
    pickup = sum(1 for r in filtered_rows if r.get("vehicle_type") == "pickup")
    sedan = sum(1 for r in filtered_rows if r.get("vehicle_type") == "sedan")
    money_totals = get_money_totals_from_records(filtered_rows, store, use_legacy_fallback=(not start and not end and not q_lower))
    iso_dates = [r.get("iso_date", "") for r in all_rows if r.get("iso_date")]
    filtered_iso_dates = [r.get("iso_date", "") for r in filtered_rows if r.get("iso_date")]

    if start or end:
        period_text = f"📊 Dashboard ช่วงวันที่ {start or '-'} ถึง {end or '-'}"
    else:
        period_text = "📊 Dashboard ข้อมูลสะสมทั้งหมด"

    return {
        "period": period_text,
        "amounts": money_totals,
        "dailyData": daily_data,
        "totals": {"motorcycle": int(motorcycle or 0), "pickup": int(pickup or 0), "sedan": int(sedan or 0), "all": int((motorcycle or 0) + (pickup or 0) + (sedan or 0))},
        "dateRange": {"start": min(iso_dates) if iso_dates else None, "end": max(iso_dates) if iso_dates else None},
        "selectedRange": {"start": min(filtered_iso_dates) if filtered_iso_dates else None, "end": max(filtered_iso_dates) if filtered_iso_dates else None},
        "recordCount": len(filtered_rows),
        "storage": "github_json",
        "app_version": APP_VERSION,
        "updated_at": store.get("updated_at"),
    }


ADMIN_HTML = """
<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vehicle Dashboard Admin</title>
<link href="https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<style>
:root{
  --bg:#eef4fb;--surface:#ffffff;--surface-2:#f8fafc;--text:#101828;--muted:#667085;
  --primary:#1d4ed8;--primary-2:#2563eb;--cyan:#12b6a6;--line:#dbe4f0;
  --shadow:0 18px 48px rgba(15,23,42,.09);--soft-shadow:0 8px 24px rgba(15,23,42,.06);
  --radius:24px;
}
*{box-sizing:border-box}
body{
  margin:0;font-family:Prompt,sans-serif;color:var(--text);letter-spacing:-.01em;
  background:radial-gradient(circle at 0% 0%,rgba(37,99,235,.16),transparent 30%),
             radial-gradient(circle at 100% 0%,rgba(20,184,166,.14),transparent 28%),
             linear-gradient(180deg,#f8fbff 0%,var(--bg) 100%);
}
a{color:inherit}button,input,textarea{font-family:Prompt,sans-serif}
.wrap{width:min(1080px,94vw);margin:0 auto;padding:22px 0 38px}
.nav{
  display:flex;align-items:center;gap:8px;margin-bottom:14px;padding:10px 12px;border-radius:20px;
  background:rgba(255,255,255,.82);border:1px solid rgba(219,228,240,.92);box-shadow:var(--soft-shadow);backdrop-filter:blur(14px)
}
.nav-brand{margin-right:auto;display:flex;align-items:center;gap:10px;font-weight:800;color:#111827}
.nav-brand .logo{width:34px;height:34px;border-radius:12px;display:grid;place-items:center;background:linear-gradient(135deg,#1d4ed8,#12b6a6);color:#fff;box-shadow:0 8px 20px rgba(37,99,235,.22)}
.nav a{text-decoration:none;color:#1d4ed8;background:#f8fbff;border:1px solid var(--line);padding:10px 14px;border-radius:14px;font-weight:800;font-size:14px;transition:.18s}
.nav a:hover{transform:translateY(-1px);background:#eff6ff}
.hero{
  position:relative;overflow:hidden;margin-bottom:16px;border-radius:28px;padding:28px 30px;
  color:#fff;background:linear-gradient(135deg,#111827 0%,#1d4ed8 58%,#12b6a6 100%);box-shadow:0 22px 58px rgba(37,99,235,.18)
}
.hero:before{content:"";position:absolute;right:-70px;top:-115px;width:270px;height:270px;border-radius:999px;background:rgba(255,255,255,.13)}
.hero-grid{position:relative;z-index:1;display:grid;grid-template-columns:minmax(0,1fr) auto;gap:16px;align-items:end}
.eyebrow{display:inline-flex;gap:8px;align-items:center;margin-bottom:10px;padding:6px 10px;border-radius:999px;background:rgba(255,255,255,.14);border:1px solid rgba(255,255,255,.2);font-size:13px;font-weight:700}
.hero h1{margin:0;font-size:clamp(30px,3.6vw,42px);line-height:1.08;letter-spacing:-.045em}
.hero p{margin:10px 0 0;color:rgba(255,255,255,.86);font-size:15px;max-width:720px}
.version{display:inline-flex;white-space:nowrap;align-items:center;padding:8px 12px;border-radius:999px;background:rgba(255,255,255,.16);border:1px solid rgba(255,255,255,.24);font-weight:800;font-size:13px}
.card{
  background:rgba(255,255,255,.92);border:1px solid rgba(219,228,240,.95);border-radius:30px;padding:22px;box-shadow:var(--shadow);backdrop-filter:blur(14px)
}
.admin-layout{display:grid;grid-template-columns:260px minmax(0,1fr);gap:20px;align-items:start}
.side-panel{background:linear-gradient(180deg,#f8fbff,#fff);border:1px solid var(--line);border-radius:24px;padding:16px;position:sticky;top:86px}
.side-title{font-size:13px;color:var(--muted);font-weight:700;margin-bottom:10px}
.tabs{display:grid;gap:8px;margin:0}
.tab-btn{width:100%;border:1px solid transparent;background:transparent;color:#475467;border-radius:16px;padding:13px 14px;text-align:left;font-weight:800;cursor:pointer;transition:.18s}
.tab-btn span{display:block;font-size:12px;font-weight:600;color:#98a2b3;margin-top:3px}
.tab-btn.active{background:#ffffff;color:#1d4ed8;border-color:#cfe0ff;box-shadow:0 10px 24px rgba(37,99,235,.10)}
.side-note{margin-top:14px;padding:12px;border-radius:16px;background:#f1f5f9;color:#667085;font-size:12px;line-height:1.55;border:1px solid #e2e8f0}
.workspace{min-width:0}
.tab-panel{display:none}
.tab-panel.active{display:block}
.panel-head{display:flex;align-items:flex-start;justify-content:space-between;gap:12px;margin-bottom:14px}
h2{margin:0;font-size:24px;letter-spacing:-.03em}
.panel-sub{margin:5px 0 0;color:var(--muted);font-size:13px}
.hint{padding:11px 13px;background:#eff6ff;color:#1d4ed8;border:1px solid #cfe0ff;border-radius:16px;margin:0 0 12px;font-size:13px;line-height:1.55}
.hint.danger{background:#fff7ed;color:#9a3412;border-color:#fed7aa}
.dropzone{
  border:1.6px dashed #93b7fb;background:linear-gradient(180deg,#ffffff,#f8fbff);border-radius:22px;padding:24px;text-align:center;
  transition:.18s;cursor:pointer;margin:12px 0;box-shadow:inset 0 1px 0 rgba(255,255,255,.8)
}
.dropzone:hover,.dropzone.dragover{border-color:#2563eb;background:#eff6ff;transform:translateY(-1px)}
.drop-icon{width:52px;height:52px;border-radius:18px;margin:0 auto 10px;display:grid;place-items:center;background:linear-gradient(135deg,#dbeafe,#ccfbf1);font-size:24px}
.drop-title{font-size:19px;font-weight:800;color:#1d4ed8}
.drop-sub{color:#667085;margin-top:5px;font-size:13px}
.file-name{display:inline-flex;margin-top:10px;padding:6px 10px;border-radius:999px;background:#f1f5f9;color:#111827;font-weight:800;font-size:12px}
input[type=file]{display:none}
textarea{width:100%;height:300px;border:1px solid var(--line);border-radius:20px;padding:14px;font-size:14px;line-height:1.65;background:#fff;color:#111827;box-shadow:inset 0 1px 2px rgba(15,23,42,.04);resize:vertical}
.row{display:flex;flex-wrap:wrap;gap:10px;margin-top:12px}
.btn{border:0;border-radius:15px;padding:12px 16px;font-weight:800;cursor:pointer;color:#fff;background:linear-gradient(135deg,#2563eb,#12b6a6);box-shadow:0 12px 26px rgba(37,99,235,.18);transition:.18s;text-decoration:none;display:inline-flex;align-items:center;justify-content:center;gap:8px}
.btn:hover{transform:translateY(-1px)}.btn:disabled{opacity:.65;cursor:not-allowed;transform:none}
.btn2{background:#f8fbff;color:#1d4ed8;border:1px solid #cfe0ff;box-shadow:none}
.status{margin-top:12px;color:var(--muted);white-space:pre-wrap;background:#f8fafc;border:1px solid var(--line);padding:13px 14px;border-radius:16px;font-size:13px;min-height:48px}
.status.loading{color:#1d4ed8;font-weight:800}
.status.loading:before{content:"";display:inline-block;width:15px;height:15px;margin-right:8px;border-radius:999px;border:3px solid #dbeafe;border-top-color:#2563eb;vertical-align:-3px;animation:spin .8s linear infinite}
@keyframes spin{to{transform:rotate(360deg)}}
@media(max-width:860px){
  .wrap{width:100%;padding:12px}
  .nav{position:relative;display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px}
  .nav-brand{grid-column:1/-1}
  .nav a{text-align:center}
  .hero{padding:24px 20px;border-radius:24px}
  .hero-grid{grid-template-columns:1fr}
  .version{width:max-content}
  .card{padding:16px;border-radius:24px}
  .admin-layout{grid-template-columns:1fr}
  .side-panel{position:relative;top:0;padding:10px;border-radius:20px}
  .tabs{grid-template-columns:1fr 1fr}
  .tab-btn{text-align:center;padding:11px 10px}
  .tab-btn span,.side-note,.side-title{display:none}
  .panel-head{display:block}
  .dropzone{padding:22px 12px}
  .row{display:grid}
  .btn{width:100%}
}
@media(max-width:520px){.tabs{grid-template-columns:1fr}.hero h1{font-size:28px}}
</style>
</head>
<body>
<div class="wrap">
  <nav class="nav">
    <div class="nav-brand"><div class="logo">🚘</div><span>Vehicle Admin Console</span></div>
    <a href="/admin">Admin</a><a href="/dashboard" target="_blank">Dashboard</a><a href="/api/health" target="_blank">Health</a>
  </nav>
  <section class="hero">
    <div class="hero-grid">
      <div>
        <div class="eyebrow">⚙️ Import Center</div>
        <h1>Vehicle Dashboard Admin</h1>
        <p>ศูนย์นำเข้าข้อมูลแบบ Replace All พร้อมตรวจสอบสถานะและส่งต่อ Dashboard สำหรับทีมปฏิบัติการ</p>
      </div>
      <div class="version">Version: v14.21 UIUX Professional Polish</div>
    </div>
  </section>
  <main class="card">
    <div class="admin-layout">
      <aside class="side-panel">
        <div class="side-title">เลือกวิธีนำเข้าข้อมูล</div>
        <div class="tabs">
          <button class="tab-btn active" data-tab="excelPanel" type="button">1) Import Excel<span>เหมาะกับไฟล์ .xlsx / .xls</span></button>
          <button class="tab-btn" data-tab="textPanel" type="button">2) Import Text<span>สำหรับข้อความรายสัปดาห์</span></button>
        </div>
        <div class="side-note">ระบบจะใช้ข้อมูลชุดล่าสุดแทนข้อมูลเดิมทั้งหมด กรุณาตรวจไฟล์ก่อนกด Upload</div>
      </aside>
      <section class="workspace">
        <section class="tab-panel active" id="excelPanel">
          <div class="panel-head"><div><h2>Import Excel</h2><p class="panel-sub">อัปโหลดไฟล์ Excel เพื่อบันทึกข้อมูลขึ้น GitHub JSON DB</p></div></div>
          <div class="hint">รองรับไฟล์ .xlsx / .xls ที่มี header: วันที่, ประเภทรถ, บริษัท, รหัส, ยอดสุทธิ, ยอดเก็บจริง</div>
          <input type="file" id="excelFile" accept=".xlsx,.xls">
          <div class="dropzone" id="excelDrop"><div class="drop-icon">⬆️</div><div class="drop-title">ลากวางไฟล์ Excel ที่นี่</div><div class="drop-sub">หรือคลิกเพื่อเลือกไฟล์จากเครื่อง</div><div class="file-name" id="excelName">ยังไม่ได้เลือกไฟล์</div></div>
          <div class="row"><button class="btn" id="excelBtn" type="button">Upload Excel และบันทึก GitHub</button><a class="btn btn2" href="/dashboard" target="_blank">เปิด Dashboard</a></div>
          <div class="status" id="excelStatus">พร้อม Import Excel</div>
        </section>
        <section class="tab-panel" id="textPanel">
          <div class="panel-head"><div><h2>Import Text</h2><p class="panel-sub">นำเข้าข้อความรายสัปดาห์หรือไฟล์ .txt</p></div></div>
          <div class="hint danger">Text Import ใช้ flow เดิมจาก v6.2 และไม่ต้องกรอก Admin Token</div>
          <input type="file" id="textFile" accept=".txt,text/plain">
          <div class="dropzone" id="textDrop"><div class="drop-icon">📄</div><div class="drop-title">ลากวางไฟล์ Text ที่นี่</div><div class="drop-sub">หรือคลิกเพื่อเลือกไฟล์ .txt จากเครื่อง / หรือวางข้อความด้านล่าง</div><div class="file-name" id="textName">ยังไม่ได้เลือกไฟล์</div></div>
          <textarea id="raw_text" placeholder="วางข้อมูลรายสัปดาห์หลายชุดต่อกันได้ตรงนี้..."></textarea>
          <div class="row"><button class="btn" id="textBtn" type="button">Import Text และบันทึก GitHub</button><a class="btn btn2" href="/dashboard" target="_blank">เปิด Dashboard</a></div>
          <div class="status" id="textStatus">พร้อม Import Text</div>
        </section>
      </section>
    </div>
  </main>
</div>
<script>

const $ = (id) => document.getElementById(id);
const excelFile = $('excelFile');
const textFile = $('textFile');
const rawText = $('raw_text');
const excelStatus = $('excelStatus');
const textStatus = $('textStatus');
const pretty = (d) => JSON.stringify(d, null, 2);

function setStatus(el, message, loading=false){
  el.classList.toggle('loading', loading);
  el.textContent = message;
}

document.querySelectorAll('.tab-btn').forEach((btn) => {
  btn.addEventListener('click', () => {
    document.querySelectorAll('.tab-btn').forEach((b) => b.classList.remove('active'));
    document.querySelectorAll('.tab-panel').forEach((p) => p.classList.remove('active'));
    btn.classList.add('active');
    $(btn.dataset.tab).classList.add('active');
  });
});

function bindDrop(zoneId, input, after){
  const zone = $(zoneId);
  zone.addEventListener('click', () => input.click());
  ['dragenter','dragover'].forEach((ev) => zone.addEventListener(ev, (e) => {
    e.preventDefault();
    zone.classList.add('dragover');
  }));
  ['dragleave','drop'].forEach((ev) => zone.addEventListener(ev, (e) => {
    e.preventDefault();
    zone.classList.remove('dragover');
  }));
  zone.addEventListener('drop', (e) => {
    const f = e.dataTransfer.files && e.dataTransfer.files[0];
    if(!f) return;
    const dt = new DataTransfer();
    dt.items.add(f);
    input.files = dt.files;
    after(f);
  });
  input.addEventListener('change', () => {
    const f = input.files && input.files[0];
    if(f) after(f);
  });
}

bindDrop('excelDrop', excelFile, (f) => { $('excelName').textContent = f.name; });
bindDrop('textDrop', textFile, async (f) => {
  $('textName').textContent = f.name;
  rawText.value = await f.text();
});

$('excelBtn').addEventListener('click', async () => {
  const file = excelFile.files && excelFile.files[0];
  if(!file){ setStatus(excelStatus, 'กรุณาเลือกไฟล์ Excel ก่อน'); return; }
  const fd = new FormData();
  fd.append('file', file);
  $('excelBtn').disabled = true;
  setStatus(excelStatus, 'กำลังอัปโหลด Excel → /api/import/excel ...', true);
  try{
    const res = await fetch('/api/import/excel', { method:'POST', body: fd });
    const data = await res.json();
    if(!res.ok){ setStatus(excelStatus, data.detail || pretty(data)); return; }
    setStatus(excelStatus, 'Import Excel สำเร็จ
' + pretty(data));
  }catch(err){
    setStatus(excelStatus, 'Import Excel ล้มเหลว: ' + err.message);
  }finally{
    $('excelBtn').disabled = false;
    excelStatus.classList.remove('loading');
  }
});

$('textBtn').addEventListener('click', async () => {
  const txt = rawText.value.trim();
  if(!txt){ setStatus(textStatus, 'กรุณาวางข้อความ หรือเลือกไฟล์ .txt ก่อน'); return; }
  const fd = new FormData();
  fd.append('raw_text', txt);
  $('textBtn').disabled = true;
  setStatus(textStatus, 'กำลัง Import Text → /api/import/text ...', true);
  try{
    const res = await fetch('/api/import/text', { method:'POST', body: fd });
    const data = await res.json();
    if(!res.ok){ setStatus(textStatus, data.detail || pretty(data)); return; }
    setStatus(textStatus, 'Import Text สำเร็จ
' + pretty(data));
  }catch(err){
    setStatus(textStatus, 'Import Text ล้มเหลว: ' + err.message);
  }finally{
    $('textBtn').disabled = false;
    textStatus.classList.remove('loading');
  }
});


</script>
</body>
</html>

"""

DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Vehicle Dashboard Only</title>
<link href="https://fonts.googleapis.com/css2?family=Prompt:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script src="https://cdn.jsdelivr.net/npm/xlsx@0.18.5/dist/xlsx.full.min.js"></script>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/flatpickr/dist/flatpickr.min.css">
<script src="https://cdn.jsdelivr.net/npm/flatpickr"></script>
<style>
:root{--bg:#f3f6fb;--card:#fff;--text:#172033;--muted:#667085;--blue:#2563eb;--green:#16a34a;--orange:#f97316;--cyan:#14b8a6;--dark:#111827;--line:#e5e7eb;--shadow:0 18px 42px rgba(15,23,42,.08);--glow:0 18px 40px rgba(37,99,235,.16);--radius:24px}
*{box-sizing:border-box}body{margin:0;font-family:Prompt,sans-serif;background:radial-gradient(circle at top left,#dbeafe 0,transparent 30%),radial-gradient(circle at top right,#ccfbf1 0,transparent 26%),linear-gradient(180deg,#f8fafc,var(--bg));color:var(--text)}.page{width:min(1280px,94vw);margin:0 auto;padding:32px 0 48px}
.hero{display:grid;grid-template-columns:1.5fr 1fr;gap:20px;margin-bottom:22px}.hero-card{background:linear-gradient(135deg,#0f172a 0%,#1d4ed8 58%,#14b8a6 100%);color:#fff;border-radius:32px;padding:32px;box-shadow:0 24px 60px rgba(37,99,235,.26);position:relative;overflow:hidden}.hero-card:before{content:"";position:absolute;width:340px;height:340px;border-radius:999px;right:-86px;top:-120px;background:rgba(255,255,255,.13)}.hero-card>*{position:relative;z-index:1}.hero-card h1{margin:0 0 10px;font-size:clamp(28px,4vw,46px);letter-spacing:-.6px}.hero-card p{opacity:.92}.period-pill{display:inline-flex;background:rgba(255,255,255,.16);border:1px solid rgba(255,255,255,.28);padding:8px 14px;border-radius:999px;margin-bottom:18px;font-weight:700;box-shadow:inset 0 1px 0 rgba(255,255,255,.24)}
.total-card,.panel,.kpi,.day-card,.toolbar,.hybrid-card{background:rgba(255,255,255,.94);box-shadow:var(--shadow);border:1px solid rgba(229,231,235,.9);backdrop-filter:blur(12px)}.total-card{border-radius:30px;padding:26px;transition:.22s}.total-card:hover,.panel:hover,.toolbar:hover{transform:translateY(-2px);box-shadow:0 22px 50px rgba(15,23,42,.11)}.label{color:var(--muted)}.amount{font-size:46px;font-weight:800;color:var(--blue);margin:8px 0}.summary-table{width:100%;border-collapse:collapse}.summary-table th,.summary-table td{padding:14px 12px;border-bottom:1px solid var(--line);text-align:left}.summary-table td:last-child,.summary-table th:last-child{text-align:right;font-weight:800}.summary-table .clickable-row{cursor:pointer;transition:.18s}.summary-table .clickable-row:hover{background:rgba(37,99,235,.06)}.row-title{display:flex;align-items:center;gap:8px;font-weight:800}.slide-caret{margin-left:auto;color:#64748b;transition:.22s}.summary-table .clickable-row.open .slide-caret{transform:rotate(180deg)}.summary-slide-row td{padding:0 12px;border-bottom:0}.summary-slide{max-height:0;overflow:hidden;opacity:0;transition:max-height .32s ease,opacity .25s ease,padding .25s ease}.summary-slide.open{max-height:220px;opacity:1;padding:10px 0 14px}.summary-subgrid{display:grid;gap:8px}.summary-subitem{display:grid;grid-template-columns:1.2fr 1fr 1fr;gap:10px;align-items:center;border:1px solid var(--line);background:rgba(248,250,252,.82);border-radius:16px;padding:11px 14px}.summary-subitem b{text-align:right}.summary-subitem .net{color:#2563eb}.summary-subitem .collected{color:#0f766e}@media(max-width:680px){.summary-subitem{grid-template-columns:1fr}.summary-subitem b{text-align:left}}
.toolbar{display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:14px;margin:0 0 16px;padding:16px;border-radius:var(--radius);transition:.22s}.filter-group{display:flex;flex-wrap:wrap;gap:10px;align-items:center}.date-input,.date-select,.search-input{border:1px solid var(--line);border-radius:14px;padding:10px 14px;font-family:Prompt,sans-serif;background:#fff;outline:none;transition:.2s}.date-input:focus,.date-select:focus,.search-input:focus{border-color:#60a5fa;box-shadow:0 0 0 4px rgba(37,99,235,.1)}.search-input{min-width:250px}.btn{border:0;border-radius:14px;padding:10px 16px;font-family:Prompt,sans-serif;font-weight:800;color:#fff;background:linear-gradient(135deg,#2563eb,#14b8a6);cursor:pointer;box-shadow:0 12px 24px rgba(37,99,235,.18);transition:.2s}.btn:hover{transform:translateY(-2px);box-shadow:0 18px 32px rgba(37,99,235,.24)}.btn2{color:#1d4ed8;background:#eff6ff;box-shadow:none}.btnDark{background:#111827}.btnToggle{background:#f8fafc;color:#1d4ed8;border:1px solid #dbeafe;box-shadow:none}.btnToggle.active{background:linear-gradient(135deg,#2563eb,#14b8a6);color:#fff;border:0;box-shadow:0 12px 24px rgba(37,99,235,.18)}
.status-pill{display:inline-flex;align-items:center;gap:8px;padding:8px 12px;border-radius:999px;background:#ecfeff;color:#155e75;font-size:13px;font-weight:800}.dot{width:8px;height:8px;border-radius:99px;background:#22c55e;box-shadow:0 0 0 5px rgba(34,197,94,.12)}.loading-overlay{position:fixed;inset:0;background:rgba(15,23,42,.28);backdrop-filter:blur(4px);display:none;align-items:center;justify-content:center;z-index:9999}.loading-overlay.show{display:flex}.loading-box{background:#fff;border-radius:24px;padding:22px 28px;box-shadow:0 28px 70px rgba(15,23,42,.22);font-weight:800;color:#172033;display:flex;gap:12px;align-items:center}.spinner{width:22px;height:22px;border-radius:999px;border:3px solid #dbeafe;border-top-color:#2563eb;animation:spin .8s linear infinite}@keyframes spin{to{transform:rotate(360deg)}}.top-toolbar{position:sticky;top:12px;z-index:20}
/* v6.11 Compact Date Range Picker */
.compact-range-toolbar{display:flex;align-items:center;justify-content:space-between;gap:16px;padding:14px 18px;border-radius:28px;position:sticky;top:12px;z-index:40}.compact-range-title h2{font-size:22px;margin:0 0 4px}.compact-range-title .date-helper{margin:0;color:var(--muted);font-size:13px}.compact-range-control{display:flex;align-items:center;gap:10px;flex-wrap:nowrap}.range-main{height:48px;min-width:300px;border:1px solid #dbeafe;border-radius:999px;background:linear-gradient(180deg,#fff,#f8fbff);padding:0 16px;display:inline-flex;align-items:center;gap:10px;font-family:Prompt,sans-serif;font-weight:800;color:#172033;cursor:pointer;box-shadow:0 10px 28px rgba(37,99,235,.08);transition:.2s}.range-main:hover{transform:translateY(-1px);border-color:#93c5fd;box-shadow:0 16px 34px rgba(37,99,235,.14)}.range-icon{width:30px;height:30px;border-radius:999px;display:inline-flex;align-items:center;justify-content:center;background:#eff6ff}.quick-actions{height:48px;display:inline-flex;align-items:center;gap:6px;padding:5px;background:#f1f5f9;border:1px solid #e5e7eb;border-radius:999px}.quick-btn{height:38px;border:0;border-radius:999px;padding:0 13px;background:transparent;color:#1d4ed8;font-family:Prompt,sans-serif;font-weight:800;cursor:pointer;white-space:nowrap}.quick-btn:hover,.quick-btn.active{background:#fff;box-shadow:0 8px 18px rgba(15,23,42,.08)}.quick-btn.muted{color:#64748b}.apply-compact{height:38px;border-radius:999px;padding:0 16px;box-shadow:none}#rangePicker{position:absolute;opacity:0;pointer-events:none;width:1px;height:1px}.flatpickr-calendar{font-family:Prompt,sans-serif;border-radius:20px!important;box-shadow:0 24px 70px rgba(15,23,42,.18)!important;border:1px solid #e5e7eb!important}.flatpickr-day.selected,.flatpickr-day.startRange,.flatpickr-day.endRange{background:#2563eb!important;border-color:#2563eb!important}.flatpickr-day.inRange{background:#dbeafe!important;border-color:#dbeafe!important;box-shadow:-5px 0 0 #dbeafe,5px 0 0 #dbeafe!important}
@media(max-width:900px){.compact-range-toolbar{display:grid;grid-template-columns:1fr;gap:12px}.compact-range-control{width:100%;display:grid;grid-template-columns:1fr}.range-main{min-width:0;width:100%;justify-content:center}.quick-actions{width:100%;height:auto;flex-wrap:wrap;border-radius:18px;justify-content:center}.apply-compact,.quick-btn{flex:1;min-width:95px}}@media(max-width:560px){.compact-range-toolbar{padding:14px}.compact-range-title h2{font-size:19px}.compact-range-title .date-helper{font-size:12px}.range-main{height:44px;font-size:13px;padding:0 12px}.quick-actions{padding:6px}.quick-btn,.apply-compact{height:36px;font-size:12px;padding:0 10px}}

.kpi-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:22px}.kpi{border-radius:var(--radius);padding:20px;transition:transform .24s ease,box-shadow .24s ease,border-color .24s ease;position:relative;overflow:hidden;animation:fadeUp .42s ease both}.kpi:nth-child(2){animation-delay:.05s}.kpi:nth-child(3){animation-delay:.1s}.kpi:nth-child(4){animation-delay:.15s}.kpi:after{content:"";position:absolute;width:120px;height:120px;border-radius:999px;right:-48px;top:-48px;background:radial-gradient(circle,rgba(37,99,235,.13),rgba(20,184,166,.04));transition:.24s}.kpi:hover{transform:translateY(-7px) scale(1.015);box-shadow:0 0 0 2px rgba(37,99,235,.1),0 24px 52px rgba(37,99,235,.18);border-color:#bfdbfe}.kpi .icon{font-size:28px;margin-bottom:8px}.kpi .value{font-size:30px;font-weight:800}.kpi .title{color:var(--muted);font-size:14px}
.section-grid{display:grid;grid-template-columns:1.2fr .8fr;gap:18px;margin-bottom:22px}.panel{border-radius:var(--radius);padding:22px;transition:.22s}.panel h2{margin:0 0 12px}.chart-wrap{height:340px}.hybrid-card{border-radius:var(--radius);padding:22px;height:100%;transition:.22s}.hybrid-card:hover{transform:translateY(-2px);box-shadow:0 22px 50px rgba(15,23,42,.11)}.hybrid-head{display:flex;align-items:flex-start;justify-content:space-between;gap:16px;margin-bottom:18px}.hybrid-total{font-size:42px;font-weight:800;color:var(--blue);line-height:1}.hybrid-label{color:var(--muted);font-size:14px;margin-top:6px}.breakdown-list{display:grid;gap:12px}.breakdown-row{display:grid;grid-template-columns:1.2fr auto;gap:12px;align-items:center;padding:13px 14px;border:1px solid #edf2f7;border-radius:18px;background:linear-gradient(180deg,#fff,#f8fafc);transition:.18s}.breakdown-row:hover{transform:translateX(4px);border-color:#bfdbfe;box-shadow:0 12px 26px rgba(37,99,235,.08)}.break-left{display:flex;align-items:center;gap:10px;font-weight:800}.break-meta{display:flex;align-items:center;gap:10px;font-weight:800}.percent{color:var(--muted);font-size:13px}.bar-track{grid-column:1/-1;height:8px;border-radius:99px;background:#eef2f7;overflow:hidden}.bar-fill{height:100%;border-radius:99px;background:linear-gradient(90deg,#2563eb,#14b8a6);width:0%;transition:width .5s}.bar-fill.orange{background:linear-gradient(90deg,#f97316,#fb923c)}.bar-fill.green{background:linear-gradient(90deg,#16a34a,#22c55e)}
.daily-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:18px}.daily-grid.compact{display:grid;grid-template-columns:1fr;gap:10px}.day-card{border-radius:var(--radius);overflow:hidden;transition:transform .25s ease,box-shadow .25s ease,border-color .25s ease,background .25s ease;animation:fadeUp .36s ease both;position:relative}.day-card:hover{transform:translateY(-7px) scale(1.01);box-shadow:0 26px 56px rgba(15,23,42,.14);border-color:#bfdbfe}.day-card.high{border-color:#93c5fd;background:linear-gradient(180deg,#eff6ff,#fff)}.day-card.low{opacity:.82}.day-card.peak:before{content:"";position:absolute;left:0;top:0;bottom:0;width:5px;background:linear-gradient(180deg,#2563eb,#14b8a6)}.day-card.compact-card{border-radius:18px}.day-head{width:100%;display:flex;align-items:center;justify-content:space-between;gap:10px;padding:18px 20px;border:0;background:linear-gradient(180deg,#fff,#fbfdff);cursor:pointer;font-family:Prompt,sans-serif;text-align:left}.compact-card .day-head{padding:14px 16px}.day-main{display:grid;gap:7px}.day-title{font-size:18px;font-weight:800}.compact-card .day-title{font-size:16px}.quick-summary{display:flex;flex-wrap:wrap;gap:8px;color:#475467;font-size:13px;font-weight:800}.mini-chip{display:inline-flex;align-items:center;gap:4px;padding:4px 9px;border-radius:999px;background:#f8fafc;border:1px solid #edf2f7}.day-tags{display:flex;gap:8px;align-items:center;justify-content:flex-end;flex-wrap:wrap}.badge{background:#eff6ff;color:#1d4ed8;border-radius:999px;padding:6px 12px;font-size:13px;font-weight:800;white-space:nowrap}.tag-peak{background:#fff7ed;color:#c2410c}.tag-low{background:#f3f4f6;color:#667085}.tag-high{background:#ecfeff;color:#0f766e}.chev{font-size:18px;color:#667085;transition:.2s}.day-card.open .chev{transform:rotate(180deg)}.day-body{max-height:0;overflow:hidden;opacity:0;transition:max-height .32s ease,opacity .25s ease,padding .25s ease;border-top:1px solid transparent;padding:0 20px}.day-card.open .day-body{max-height:900px;opacity:1;padding:0 20px 20px;border-top-color:var(--line)}.vehicle-group{margin-top:14px}.vehicle-title{font-weight:800;margin-bottom:8px}.company{display:inline-flex;margin:6px 0 4px;padding:4px 10px;border-radius:999px;background:#f3f4f6;font-size:12px;font-weight:800}ul{list-style:none;padding:0;margin:0;display:grid;gap:7px}li{background:#f9fafb;border:1px solid #eef2f7;border-radius:14px;padding:9px 11px;font-size:13px;transition:.18s}li:hover{background:#eff6ff;border-color:#bfdbfe;transform:translateX(3px)}.day-money-summary{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:10px;margin:16px 0 4px}.money-pill{border-radius:16px;padding:12px 14px;display:grid;gap:3px}.money-pill.net{border:1px solid #bfdbfe;background:linear-gradient(135deg,#eff6ff,#ffffff)}.money-pill.collected{border:1px solid #99f6e4;background:linear-gradient(135deg,#ecfeff,#f0fdfa)}.money-pill span,.money-pill small{color:#64748b;font-size:12px;font-weight:800}.money-pill strong{font-size:20px;color:#0f172a}.money-pill.net strong{color:#1d4ed8}.money-pill.collected strong{color:#0f766e}.vehicle-title-row{display:flex;align-items:center;justify-content:space-between;gap:12px;flex-wrap:wrap}.vehicle-money{display:flex;gap:6px;flex-wrap:wrap;justify-content:flex-end}.money-chip{font-size:12px;font-weight:800;border-radius:999px;padding:5px 10px;white-space:nowrap}.money-chip.net{color:#1d4ed8;background:#eff6ff;border:1px solid #bfdbfe}.money-chip.collected{color:#0f766e;background:#ecfeff;border:1px solid #99f6e4}.vehicle-item-row{display:flex;align-items:center;justify-content:space-between;gap:10px;flex-wrap:wrap}.item-money{display:flex;gap:6px;flex-wrap:wrap;justify-content:flex-end}.item-money span{font-size:12px;font-weight:800;border-radius:999px;padding:4px 8px}.item-money .net{color:#1d4ed8;background:#eff6ff;border:1px solid #bfdbfe}.item-money .collected{color:#0f766e;background:#ecfeff;border:1px solid #99f6e4}@media(max-width:680px){.day-money-summary{grid-template-columns:1fr}.vehicle-item-row{align-items:flex-start}.item-money,.vehicle-money{justify-content:flex-start}}.pagination{display:flex;flex-wrap:wrap;gap:10px;align-items:center;justify-content:center;margin:20px 0}.page-info{color:var(--muted);font-weight:700}
@keyframes fadeUp{from{opacity:0;transform:translateY(12px)}to{opacity:1;transform:translateY(0)}}.company-kpi .company-dot{width:12px;height:12px;border-radius:999px;display:inline-block;margin-right:6px}.company-kpi .company-meta{display:flex;justify-content:space-between;margin-top:10px;color:var(--muted);font-size:13px;font-weight:800}.company-kpi .company-progress{height:8px;border-radius:999px;background:#eef2f7;overflow:hidden;margin-top:8px}.company-kpi .company-progress span{display:block;height:100%;border-radius:999px;width:0%;transition:width .45s ease}.company-kpi[data-company="RVP"]{border-left:5px solid #2563eb}.company-kpi[data-company="ERGO"]{border-left:5px solid #dc2626}.company-kpi[data-company="TPB"]{border-left:5px solid #0ea5e9}.company-kpi[data-company="TOTAL"]{border-left:5px solid #111827}.company-detail{margin-top:14px;padding-top:12px;border-top:1px solid #eef2f7;display:grid;gap:7px;color:#475467;font-size:12px;font-weight:700}.company-detail .line{display:flex;justify-content:space-between;gap:10px}.company-detail b{color:#172033}.company-kpi.active{transform:translateY(-8px) scale(1.025);box-shadow:0 0 0 3px rgba(37,99,235,.12),0 28px 60px rgba(37,99,235,.2);border-color:#bfdbfe}.company-kpi.dim{opacity:.58}
@media(max-width:980px){.hero,.section-grid,.daily-grid{grid-template-columns:1fr}.kpi-grid{grid-template-columns:repeat(2,1fr)}}@media(max-width:560px){.kpi-grid{grid-template-columns:1fr}.filter-group,.date-input,.date-select,.search-input,.btn{width:100%}.amount,.hybrid-total{font-size:38px}.day-head{align-items:flex-start}.day-tags{justify-content:flex-start}}

/* v6.8.1 layout/error hotfix */
.hero{grid-template-columns:.95fr 1fr;align-items:stretch}.top-toolbar{padding:12px 16px;border-radius:22px;display:flex;align-items:center;justify-content:space-between;gap:16px}.top-toolbar h2{font-size:22px;margin:0;white-space:nowrap}.top-toolbar .filter-group{justify-content:flex-end}.top-toolbar .date-input{width:160px;height:42px;padding:8px 12px}.top-toolbar .btn{height:42px;padding:0 16px}.toolbar .search-input{min-width:210px;max-width:280px}.dashboard-version{display:none}.version-credit{position:fixed;right:14px;bottom:10px;font-size:10px;color:#98a2b3;background:rgba(255,255,255,.72);border:1px solid #eef2f7;border-radius:999px;padding:4px 8px;z-index:50;backdrop-filter:blur(6px)}.hybrid-card.primary{min-height:100%;display:flex;flex-direction:column;justify-content:space-between}.hybrid-card.primary .period-pill{color:#1d4ed8;background:#eff6ff;border-color:#bfdbfe;margin-bottom:12px}.hybrid-card.primary h2{font-size:28px;margin:0 0 8px}.total-card{min-height:100%}.section-grid{display:none!important}
@media(max-width:980px){.hero{grid-template-columns:1fr}.top-toolbar{position:relative;top:auto;display:block}.top-toolbar .filter-group{justify-content:flex-start}.toolbar .search-input{max-width:none}}
@media(max-width:560px){.top-toolbar{padding:14px}.top-toolbar h2{font-size:20px;margin-bottom:12px}.top-toolbar .filter-group{display:grid;grid-template-columns:1fr;gap:8px;width:100%}.top-toolbar .date-input{width:100%;height:38px;font-size:12px;padding:7px 9px}.top-toolbar .btn{width:100%;height:38px;padding:0 12px}.hero{gap:14px}.hybrid-card.primary h2{font-size:25px}.amount,.hybrid-total{font-size:38px}}
</style></head>
<body><div class="loading-overlay" id="loadingOverlay"><div class="loading-box"><span class="spinner"></span><span id="loadingText">กำลังโหลดข้อมูล...</span></div></div><main class="page">
<section class="toolbar top-toolbar compact-range-toolbar"><div class="compact-range-title"><h2>เลือกช่วงวันที่ Dashboard</h2><p class="date-helper">เลือกช่วงวันที่แบบใหม่ มีผลกับข้อมูลทั้งหน้า Dashboard</p></div><div class="compact-range-control"><button class="range-main" id="rangeOpenBtn" type="button"><span class="range-icon">📅</span><span id="rangeDisplayText">01 May 2026 → 31 May 2026</span></button><input id="rangePicker" type="text" aria-label="เลือกช่วงวันที่" autocomplete="off"><input id="startDate" type="hidden"><input id="endDate" type="hidden"><div class="quick-actions"><button class="quick-btn" id="todayBtn" type="button">วันนี้</button><button class="quick-btn" id="thisWeekBtn" type="button">สัปดาห์นี้</button><button class="quick-btn active" id="thisMonthBtn" type="button">เดือนนี้</button><button class="btn apply-compact" id="applyBtn" type="button">แสดงผล</button><button class="quick-btn muted" id="resetBtn" type="button">ดูทั้งหมด</button></div></div></section>
<section class="hero"><div class="hybrid-card primary"><div><div class="period-pill" id="period">📊 Dashboard ข้อมูลสะสมทั้งหมด</div><h2>สัดส่วนประเภทรถ</h2><p class="label" style="margin:0 0 14px">ภาพรวมจำนวนรถตามช่วงวันที่ที่เลือก</p></div><div><div class="hybrid-head"><div><div class="hybrid-total" id="hybridTotal">0</div><div class="hybrid-label">คันทั้งหมด</div></div><div><span class="status-pill"><span class="dot"></span><span id="refreshStatus">Auto refresh ทุก 15 นาที</span></span></div></div><div class="breakdown-list" id="breakdownList"></div></div></div><div class="total-card"><div class="label">ยอดเก็บจริง</div><div class="amount" id="netTotalAmount">0</div><div class="label">ยอดสุทธิตามระบบ</div><div class="amount" id="collectedTotalAmount" style="font-size:34px;color:#16a34a;margin-top:10px">0</div><table class="summary-table"><tr><th>หมวด</th><th>ยอดสุทธิ</th><th>ยอดเก็บจริง</th></tr><tr class="clickable-row" id="carSummaryToggle" onclick="toggleCarSummary()"><td><span class="row-title">🚛 🚗 รถยนต์รวม <span class="slide-caret">⌄</span></span></td><td id="carNetAmount">0 บาท</td><td id="carCollectedAmount">0 บาท</td></tr><tr class="summary-slide-row"><td colspan="3"><div class="summary-slide" id="carSummarySlide"><div class="summary-subgrid"><div class="summary-subitem"><span>🚛 รถกระบะ</span><b class="net" id="pickupNetAmount">0 บาท</b><b class="collected" id="pickupCollectedAmount">0 บาท</b></div><div class="summary-subitem"><span>🚗 รถยนต์เก๋ง</span><b class="net" id="sedanNetAmount">0 บาท</b><b class="collected" id="sedanCollectedAmount">0 บาท</b></div></div></div></td></tr><tr><td>🏍 รถจักรยานยนต์</td><td id="motorNetAmount">0 บาท</td><td id="motorCollectedAmount">0 บาท</td></tr></table></div></section>
<section class="kpi-grid">
 <div class="kpi company-kpi" data-company="RVP" onmouseenter="highlightCompany('RVP')" onmouseleave="highlightCompany(null)"><div class="icon"><span class="company-dot" style="background:#2563eb"></span>บริษัทกลางฯ RVP</div><div class="value" id="rvpCount">0</div><div class="title">จำนวนรถทั้งหมด</div><div class="company-meta"><span>Share</span><span id="rvpPercent">0%</span></div><div class="company-progress"><span id="rvpBar" style="background:linear-gradient(90deg,#2563eb,#1d4ed8)"></span></div><div class="company-detail" id="rvpDetail"></div></div>
 <div class="kpi company-kpi" data-company="ERGO" onmouseenter="highlightCompany('ERGO')" onmouseleave="highlightCompany(null)"><div class="icon"><span class="company-dot" style="background:#dc2626"></span>ERGO</div><div class="value" id="ergoCount">0</div><div class="title">จำนวนรถทั้งหมด</div><div class="company-meta"><span>Share</span><span id="ergoPercent">0%</span></div><div class="company-progress"><span id="ergoBar" style="background:linear-gradient(90deg,#dc2626,#ef4444)"></span></div><div class="company-detail" id="ergoDetail"></div></div>
 <div class="kpi company-kpi" data-company="TPB" onmouseenter="highlightCompany('TPB')" onmouseleave="highlightCompany(null)"><div class="icon"><span class="company-dot" style="background:#0ea5e9"></span>ไทยไพบูลย์ TPB</div><div class="value" id="tpbCount">0</div><div class="title">จำนวนรถทั้งหมด</div><div class="company-meta"><span>Share</span><span id="tpbPercent">0%</span></div><div class="company-progress"><span id="tpbBar" style="background:linear-gradient(90deg,#0ea5e9,#06b6d4)"></span></div><div class="company-detail" id="tpbDetail"></div></div>
 <div class="kpi company-kpi" data-company="TOTAL" onmouseenter="highlightCompany('TOTAL')" onmouseleave="highlightCompany(null)"><div class="icon"><span class="company-dot" style="background:#111827"></span>รวม</div><div class="value" id="companyTotalCount">0</div><div class="title">จำนวนรถรวมทั้งหมด</div><div class="company-meta"><span>Share</span><span>100%</span></div><div class="company-progress"><span style="width:100%;background:linear-gradient(90deg,#111827,#64748b)"></span></div><div class="company-detail" id="totalDetail"></div></div>
</section>

<section class="toolbar"><h2>รายการแยกรายวัน</h2><div class="filter-group"><input class="search-input" id="searchBox" placeholder="ค้นหาทะเบียน / เลขกรมธรรม์ / บริษัท"><select class="date-select" id="dateFilter"><option value="all">ดูทั้งหมด</option></select><button class="btn" id="showDateBtn">แสดงวันที่เลือก</button><button class="btn btn2" id="showAllBtn">ดูทั้งหมด</button><button class="btn btnToggle active" id="detailModeBtn">📄 Detail</button><button class="btn btnToggle" id="compactModeBtn">⚡ Compact</button><button class="btn btnDark" id="exportPdfBtn">Export PDF</button><button class="btn btnDark" id="exportExcelBtn">Export Excel</button></div></section>
<section class="daily-grid" id="cards"></section><div class="pagination"><button class="btn btn2" id="prevPageBtn">ก่อนหน้า</button><span class="page-info" id="pageInfo">Page 1</span><button class="btn btn2" id="nextPageBtn">ถัดไป</button></div><p class="status-pill" id="status">Loading...</p>
<div class="version-credit">Version: __APP_VERSION__</div>
</main>
<script>
let report=null,allDays=[],filteredDays=[],viewDays=[],dailyChart=null,rangePickerInstance=null;let currentPage=1,pageSize=8,viewMode='detail',activeSelected='all';const box=id=>document.getElementById(id);const money=n=>Math.round(n||0).toLocaleString('th-TH');function destroy(){if(dailyChart)dailyChart.destroy()}function fmtDate(d){const y=d.getFullYear(),m=String(d.getMonth()+1).padStart(2,'0'),day=String(d.getDate()).padStart(2,'0');return `${y}-${m}-${day}`}function parseISODate(v){const parts=String(v||'').split('-').map(Number);return parts.length===3?new Date(parts[0],parts[1]-1,parts[2]):new Date()}function prettyRange(s,e){if(!s&&!e)return 'ข้อมูลสะสมทั้งหมด';const opt={day:'2-digit',month:'short',year:'numeric'};const a=s?parseISODate(s).toLocaleDateString('en-GB',opt):'เริ่มต้น';const b=e?parseISODate(e).toLocaleDateString('en-GB',opt):'สิ้นสุด';return `${a} → ${b}`}function updateRangeLabel(){const s=box('startDate')?.value||'',e=box('endDate')?.value||'';if(box('rangeDisplayText'))box('rangeDisplayText').textContent=prettyRange(s,e);if(rangePickerInstance&&s&&e)rangePickerInstance.setDate([s,e],false,'Y-m-d')}function setActiveQuick(id){['todayBtn','thisWeekBtn','thisMonthBtn'].forEach(x=>box(x)?.classList.remove('active'));if(id)box(id)?.classList.add('active')}function setRange(start,end,quickId){box('startDate').value=start||'';box('endDate').value=end||'';setActiveQuick(quickId);updateRangeLabel()}function setCurrentMonthDefault(){const now=new Date();setRange(fmtDate(new Date(now.getFullYear(),now.getMonth(),1)),fmtDate(new Date(now.getFullYear(),now.getMonth()+1,0)),'thisMonthBtn')}function setToday(){const now=new Date();setRange(fmtDate(now),fmtDate(now),'todayBtn')}function setThisWeek(){const now=new Date();const day=now.getDay();const diffToMonday=(day===0?-6:1-day);const start=new Date(now);start.setDate(now.getDate()+diffToMonday);const end=new Date(start);end.setDate(start.getDate()+6);setRange(fmtDate(start),fmtDate(end),'thisWeekBtn')}function setupRange(){const dates=allDays.map(d=>d.isoDate).filter(Boolean).sort();if(!box('startDate').value)box('startDate').value=dates[0]||'';if(!box('endDate').value)box('endDate').value=dates[dates.length-1]||'';updateRangeLabel()}function itemText(item){return (item&&typeof item==='object')?(item.text||''):String(item||'')}function itemNet(item){return Number((item&&typeof item==='object')?(item.net_amount||0):0)||0}function itemCollected(item){return Number((item&&typeof item==='object')?(item.collected_amount||0):0)||0}function flattenRows(days){const rows=[];days.forEach(day=>day.groups.forEach(g=>g.items.forEach(item=>rows.push({date:day.date,type:g.title,company:g.company||'',item:itemText(item),net_amount:itemNet(item),collected_amount:itemCollected(item)}))));return rows}
function animateNumber(el,target){const end=Number(target)||0;const start=Number((el.textContent||'0').replace(/,/g,''))||0;const duration=420;const t0=performance.now();function tick(now){const p=Math.min(1,(now-t0)/duration);const eased=1-Math.pow(1-p,3);el.textContent=money(start+(end-start)*eased);if(p<1)requestAnimationFrame(tick);else el.textContent=money(end)}requestAnimationFrame(tick)}
function colorWithAlpha(hex,alpha){const map={'#2563eb':'37,99,235','#dc2626':'220,38,38','#0ea5e9':'14,165,233','#111827':'17,24,39','#94a3b8':'148,163,184'};return `rgba(${map[hex]||'37,99,235'},${alpha})`}function applyChartHighlight(index){if(!dailyChart)return;const colors=['#2563eb','#dc2626','#0ea5e9'];dailyChart.data.datasets.forEach((ds,di)=>{if(ds.type==='line'){ds.borderColor=index==null?'#111827':colorWithAlpha('#111827',.95);ds.backgroundColor=ds.borderColor;ds.pointBackgroundColor=ds.data.map((_,i)=>index==null||i===index?'#111827':colorWithAlpha('#111827',.18));return}ds.backgroundColor=ds.data.map((_,i)=>index==null||i===index?colors[di]:colorWithAlpha(colors[di],.18))});dailyChart.update('none')}
function getCompanyData(){
 return filteredDays.map(day=>{
   let RVP=0, ERGO=0, TPB=0, UNKNOWN=0;
   (day.groups||[]).forEach(g=>{
     const company=(g.company||'').toLowerCase();
     const count=(g.items&&g.items.length)?g.items.length:(g.count||0);
     if(g.key==='motorcycle'){
       RVP += count;
     }else if(company.includes('ergo')){
       ERGO += count;
     }else if(company.includes('ไทยไพบูลย์') || company.includes('tpb')){
       TPB += count;
     }else{
       UNKNOWN += count;
     }
   });
   return {date:day.date,label:day.date.slice(0,5),RVP,ERGO,TPB,UNKNOWN,total:RVP+ERGO+TPB+UNKNOWN};
 });
}

function getCompanySummary(){
 const companyData=getCompanyData();
 return companyData.reduce((acc,d)=>{acc.RVP+=d.RVP;acc.ERGO+=d.ERGO;acc.TPB+=d.TPB;acc.UNKNOWN+=(d.UNKNOWN||0);acc.total+=d.total;return acc;},{RVP:0,ERGO:0,TPB:0,UNKNOWN:0,total:0});
}
function companyKeyFromName(name,key){const c=String(name||'').toLowerCase();if(key==='motorcycle'||c.includes('rvp')||c.includes('บริษัทกลาง'))return 'RVP';if(c.includes('ergo'))return 'ERGO';if(c.includes('ไทยไพบูลย์')||c.includes('tpb'))return 'TPB';return 'UNKNOWN'}
function emptyCompanySummary(){return {count:0,net:0,collected:0,types:{motorcycle:0,pickup:0,sedan:0,other:0}}}
function getCompanyDetailedSummary(){const out={RVP:emptyCompanySummary(),ERGO:emptyCompanySummary(),TPB:emptyCompanySummary(),TOTAL:emptyCompanySummary(),UNKNOWN:emptyCompanySummary()};filteredDays.forEach(day=>{(day.groups||[]).forEach(g=>{const key=companyKeyFromName(g.company,g.key);(g.items||[]).forEach(item=>{const target=out[key]||out.UNKNOWN;const type=(item&&typeof item==='object'&&item.vehicle_type)||g.key||'other';target.count++;target.net+=itemNet(item);target.collected+=itemCollected(item);if(target.types[type]!==undefined)target.types[type]++;else target.types.other++;out.TOTAL.count++;out.TOTAL.net+=itemNet(item);out.TOTAL.collected+=itemCollected(item);if(out.TOTAL.types[type]!==undefined)out.TOTAL.types[type]++;else out.TOTAL.types.other++;})})});return out}
function companyDetailHtml(d){return `<div class="line"><span>🏍 มอเตอร์ไซค์</span><b>${d.types.motorcycle||0} คัน</b></div><div class="line"><span>🚛 กระบะ</span><b>${d.types.pickup||0} คัน</b></div><div class="line"><span>🚗 เก๋ง</span><b>${d.types.sedan||0} คัน</b></div><div class="line"><span>ยอดสุทธิตามระบบ</span><b>${money(d.net)} บาท</b></div><div class="line"><span>ยอดเก็บจริง</span><b>${money(d.collected)} บาท</b></div>`}
function setCompanyKPI(){
 const detail=getCompanyDetailedSummary();
 const s={RVP:detail.RVP.count,ERGO:detail.ERGO.count,TPB:detail.TPB.count,total:detail.TOTAL.count};
 const pct=(v)=>s.total?Math.round((v/s.total)*100):0;
 animateNumber(box('rvpCount'),s.RVP);animateNumber(box('ergoCount'),s.ERGO);animateNumber(box('tpbCount'),s.TPB);animateNumber(box('companyTotalCount'),s.total);
 box('rvpPercent').textContent=pct(s.RVP)+'%';box('ergoPercent').textContent=pct(s.ERGO)+'%';box('tpbPercent').textContent=pct(s.TPB)+'%';
 box('rvpBar').style.width=pct(s.RVP)+'%';box('ergoBar').style.width=pct(s.ERGO)+'%';box('tpbBar').style.width=pct(s.TPB)+'%';
 if(box('rvpDetail'))box('rvpDetail').innerHTML=companyDetailHtml(detail.RVP);if(box('ergoDetail'))box('ergoDetail').innerHTML=companyDetailHtml(detail.ERGO);if(box('tpbDetail'))box('tpbDetail').innerHTML=companyDetailHtml(detail.TPB);if(box('totalDetail'))box('totalDetail').innerHTML=companyDetailHtml(detail.TOTAL);
}
function colorWithAlpha(hex,alpha){const map={'#2563eb':'37,99,235','#dc2626':'220,38,38','#0ea5e9':'14,165,233','#111827':'17,24,39','#94a3b8':'148,163,184'};return `rgba(${map[hex]||'37,99,235'},${alpha})`}
function setCompanyCardsState(company){document.querySelectorAll('.company-kpi').forEach(card=>{const c=card.dataset.company;card.classList.toggle('active',!!company&&c===company);card.classList.toggle('dim',!!company&&c!==company&&company!=='TOTAL')})}
function highlightCompany(company){
 setCompanyCardsState(company);
 if(!dailyChart)return;
 const colorMap={RVP:'#2563eb',ERGO:'#dc2626',TPB:'#0ea5e9',UNKNOWN:'#94a3b8'};
 dailyChart.data.datasets.forEach(ds=>{
   const label=ds.label||'';
   if(ds.type==='line'){const active=!company||company==='TOTAL';ds.borderColor=active?'#111827':colorWithAlpha('#111827',.18);ds.backgroundColor=ds.borderColor;ds.pointBackgroundColor=ds.data.map(()=>active?'#111827':colorWithAlpha('#111827',.18));return}
   let key='UNKNOWN';if(label.includes('RVP'))key='RVP';if(label.includes('ERGO'))key='ERGO';if(label.includes('TPB'))key='TPB';
   const active=!company||company==='TOTAL'||company===key;const base=colorMap[key]||'#94a3b8';ds.backgroundColor=ds.data.map(()=>active?base:colorWithAlpha(base,.16));
 });
 dailyChart.update('none');
}
function renderCharts(){return}
function renderBreakdown(motor,pickup,sedan,total){box('hybridTotal').textContent=money(total);const rows=[{icon:'🏍',label:'รถจักรยานยนต์',value:motor,cls:''},{icon:'🚛',label:'รถกระบะ',value:pickup,cls:'orange'},{icon:'🚗',label:'รถยนต์เก๋ง',value:sedan,cls:'green'}];box('breakdownList').innerHTML=rows.map(r=>{const pct=total?Math.round((r.value/total)*100):0;return `<div class="breakdown-row"><div class="break-left"><span>${r.icon}</span><span>${r.label}</span></div><div class="break-meta"><span>${money(r.value)}</span><span class="percent">${pct}%</span></div><div class="bar-track"><div class="bar-fill ${r.cls}" style="width:${pct}%"></div></div></div>`}).join('')}
function toggleCarSummary(){const slide=box('carSummarySlide');const row=box('carSummaryToggle');if(!slide)return;slide.classList.toggle('open');row?.classList.toggle('open',slide.classList.contains('open'))}
function render(selected='all'){
 const t=report.totals||{};
 const motor=t.motorcycle||0,pickup=t.pickup||0,sedan=t.sedan||0,total=t.all||0;
 const net=report.amounts?.net||{};
 const collected=report.amounts?.collected||{};
 box('period').textContent=report.period||'📊 Dashboard ข้อมูลสะสมทั้งหมด';
 box('netTotalAmount').textContent=money(collected.total)+' บาท';
 box('collectedTotalAmount').textContent=money(net.total)+' บาท';
 box('pickupNetAmount').textContent=money(net.pickup)+' บาท';
 box('pickupCollectedAmount').textContent=money(collected.pickup)+' บาท';
 box('sedanNetAmount').textContent=money(net.sedan)+' บาท';
 box('sedanCollectedAmount').textContent=money(collected.sedan)+' บาท';
 box('carNetAmount').textContent=money(net.car)+' บาท';
 box('carCollectedAmount').textContent=money(collected.car)+' บาท';
 box('motorNetAmount').textContent=money(net.motorcycle)+' บาท';
 box('motorCollectedAmount').textContent=money(collected.motorcycle)+' บาท';
 box('dateFilter').innerHTML='<option value="all">ดูทั้งหมด</option>'+filteredDays.map(d=>`<option value="${d.date}">${d.date}</option>`).join('');
 renderCharts();
 setCompanyKPI();
 renderBreakdown(motor,pickup,sedan,total);
 currentPage=1;
 renderCards(selected);
 const sr=report.selectedRange||{};const rangeText=(sr.start&&sr.end)?`ช่วงวันที่ ${sr.start} ถึง ${sr.end}`:'ข้อมูลสะสมทั้งหมด';box('status').textContent=`${rangeText} • จำนวนรถ ${total} คัน • แสดง ${filteredDays.length}/${allDays.length} วัน`;
}
function getCardList(selected='all'){const base=selected==='all'?filteredDays:filteredDays.filter(d=>d.date===selected);const q=box('searchBox').value.trim().toLowerCase();if(!q)return base;return base.map(day=>{const groups=day.groups.map(g=>{const items=g.items.filter(i=>(day.date+' '+g.title+' '+(g.company||'')+' '+itemText(i)).toLowerCase().includes(q));return {...g,items,count:items.length}}).filter(g=>g.items.length);return {...day,groups,motorcycle:groups.filter(g=>g.key==='motorcycle').reduce((s,g)=>s+g.items.length,0),pickup:groups.filter(g=>g.key==='pickup').reduce((s,g)=>s+g.items.length,0),sedan:groups.filter(g=>g.key==='sedan').reduce((s,g)=>s+g.items.length,0)}}).filter(d=>d.groups.length)}
function getDayMeta(list,day){const totals=list.map(d=>d.motorcycle+d.pickup+d.sedan);const max=Math.max(...totals,0),min=Math.min(...totals,0);const total=day.motorcycle+day.pickup+day.sedan;let tags=[],cls=[];if(total===max&&max>0){tags.push('🔥 Peak');cls.push('peak','high')}else if(total>=max*.75&&max>0){tags.push('เด่น');cls.push('high')}if(total===min&&list.length>1){tags.push('Low');cls.push('low')}return {total,tags,cls:cls.join(' ')}}
function sumItems(items){return (items||[]).reduce((acc,item)=>{acc.net+=itemNet(item);acc.collected+=itemCollected(item);return acc},{net:0,collected:0})}
function sumDay(day){return (day.groups||[]).reduce((acc,g)=>{const s=sumItems(g.items);acc.net+=s.net;acc.collected+=s.collected;return acc},{net:0,collected:0})}
function amountLine(net,collected){return `<div class="item-money"><span class="net">ยอดสุทธิ ${money(net)} บาท</span><span class="collected">ยอดเก็บจริง ${money(collected)} บาท</span></div>`}
function buildDetails(day){const daySum=sumDay(day);return `<div class="day-money-summary"><div class="money-pill net"><span>ยอดสุทธิตามระบบ</span><strong>${money(daySum.net)} บาท</strong><small>รวมทุกคัน</small></div><div class="money-pill collected"><span>ยอดเงินรวมประจำวัน</span><strong>${money(daySum.collected)} บาท</strong><small>ยอดเก็บจริง</small></div></div>`+day.groups.map(g=>{const gSum=sumItems(g.items);return `<div class="vehicle-group"><div class="vehicle-title vehicle-title-row"><span>${g.icon} ${g.title} (${g.items.length} คัน)</span><span class="vehicle-money"><span class="money-chip net">สุทธิ ${money(gSum.net)} บาท</span><span class="money-chip collected">เก็บจริง ${money(gSum.collected)} บาท</span></span></div>${g.company?`<span class="company">${g.company}</span>`:''}<ul>${g.items.map(i=>`<li><div class="vehicle-item-row"><span>${itemText(i)}</span>${amountLine(itemNet(i),itemCollected(i))}</div></li>`).join('')}</ul></div>`}).join('')}
function toggleDay(btn,index){const card=btn.closest('.day-card');const body=card.querySelector('.day-body');if(card.classList.contains('open')){card.classList.remove('open');return}if(!body.dataset.loaded){const day=viewDays[index];body.innerHTML=buildDetails(day);body.dataset.loaded='1'}card.classList.add('open')}
function renderCards(selected='all'){activeSelected=selected;const list=getCardList(selected);viewDays=list;const totalPages=Math.max(1,Math.ceil(list.length/pageSize));if(currentPage>totalPages)currentPage=totalPages;const start=(currentPage-1)*pageSize;const pageItems=list.slice(start,start+pageSize);box('cards').classList.toggle('compact',viewMode==='compact');box('pageInfo').textContent=`หน้า ${currentPage}/${totalPages} • แสดง ${pageItems.length}/${list.length} วัน`;box('prevPageBtn').disabled=currentPage<=1;box('nextPageBtn').disabled=currentPage>=totalPages;if(!pageItems.length){box('cards').innerHTML='<article class="day-card"><button class="day-head"><span class="day-title">ไม่พบข้อมูล</span></button></article>';return}box('cards').innerHTML=pageItems.map((day,idx)=>{const globalIndex=start+idx;const meta=getDayMeta(list,day);const compact=viewMode==='compact';const open=!compact&&idx<2;const tags=meta.tags.map(t=>`<span class="badge ${t.includes('Peak')?'tag-peak':t.includes('Low')?'tag-low':'tag-high'}">${t}</span>`).join('');const summary=`<div class="quick-summary"><span class="mini-chip">🏍 ${day.motorcycle}</span><span class="mini-chip">🚛 ${day.pickup}</span><span class="mini-chip">🚗 ${day.sedan}</span></div>`;const bodyContent=open?buildDetails(day):'';return `<article class="day-card ${meta.cls} ${compact?'compact-card':''} ${open?'open':''}"><button class="day-head" onclick="toggleDay(this,${globalIndex})"><span class="day-main"><span class="day-title">📊 วันที่ ${day.date}</span>${summary}</span><span class="day-tags">${tags}<span class="badge">รวม ${meta.total} คัน</span><span class="chev">⌄</span></span></button><div class="day-body" data-loaded="${open?'1':''}">${bodyContent}</div></article>`}).join('')}
function exportExcel(){const safeText=(id,fallback='0')=>box(id)?.textContent?.replace(' บาท','')||fallback;const rows=flattenRows(viewDays.length?viewDays:filteredDays);const sourceDays=(viewDays.length?viewDays:filteredDays)||[];const counts=sourceDays.reduce((acc,day)=>{acc.motorcycle+=Number(day.motorcycle||0);acc.pickup+=Number(day.pickup||0);acc.sedan+=Number(day.sedan||0);return acc},{motorcycle:0,pickup:0,sedan:0});counts.all=counts.motorcycle+counts.pickup+counts.sedan;const summaryRows=[{หมวด:'ยอดเก็บจริงรวมทั้งหมด',ยอด:safeText('netTotalAmount'),หน่วย:'บาท'},{หมวด:'ยอดสุทธิตามระบบรวมทั้งหมด',ยอด:safeText('collectedTotalAmount'),หน่วย:'บาท'},{หมวด:'รถกระบะ - ยอดสุทธิ',ยอด:safeText('pickupNetAmount'),หน่วย:'บาท'},{หมวด:'รถกระบะ - ยอดเก็บจริง',ยอด:safeText('pickupCollectedAmount'),หน่วย:'บาท'},{หมวด:'รถยนต์เก๋ง - ยอดสุทธิ',ยอด:safeText('sedanNetAmount'),หน่วย:'บาท'},{หมวด:'รถยนต์เก๋ง - ยอดเก็บจริง',ยอด:safeText('sedanCollectedAmount'),หน่วย:'บาท'},{หมวด:'รถยนต์รวม - ยอดสุทธิ',ยอด:safeText('carNetAmount'),หน่วย:'บาท'},{หมวด:'รถยนต์รวม - ยอดเก็บจริง',ยอด:safeText('carCollectedAmount'),หน่วย:'บาท'},{หมวด:'รถจักรยานยนต์ - ยอดสุทธิ',ยอด:safeText('motorNetAmount'),หน่วย:'บาท'},{หมวด:'รถจักรยานยนต์ - ยอดเก็บจริง',ยอด:safeText('motorCollectedAmount'),หน่วย:'บาท'},{หมวด:'รถจักรยานยนต์',ยอด:counts.motorcycle,หน่วย:'คัน'},{หมวด:'รถกระบะ',ยอด:counts.pickup,หน่วย:'คัน'},{หมวด:'รถยนต์เก๋ง',ยอด:counts.sedan,หน่วย:'คัน'},{หมวด:'จำนวนรถรวมทั้งหมด',ยอด:counts.all,หน่วย:'คัน'}];const wsSummary=XLSX.utils.json_to_sheet(summaryRows);const wsDetail=XLSX.utils.json_to_sheet(rows.map(r=>({วันที่:r.date,ประเภทรถ:r.type,บริษัท:r.company,รายการ:r.item,ยอดสุทธิตามระบบ:r.net_amount,ยอดเก็บจริง:r.collected_amount})));const wb=XLSX.utils.book_new();XLSX.utils.book_append_sheet(wb,wsSummary,'Summary');XLSX.utils.book_append_sheet(wb,wsDetail,'Detail');XLSX.writeFile(wb,'vehicle-dashboard.xlsx')}
function escapeHtml(text){return String(text||'').replaceAll('&','&amp;').replaceAll('<','&lt;').replaceAll('>','&gt;').replaceAll('"','&quot;').replaceAll("'","&#039;")}
function exportPDF(){
 const rows=flattenRows(viewDays.length?viewDays:filteredDays);
 const printedAt=new Date().toLocaleString('th-TH');
 const totalAmount=box('netTotalAmount')?.textContent||'0';
 const netAmount=box('collectedTotalAmount')?.textContent||'0';
 const html=`<!DOCTYPE html><html lang="th"><head><meta charset="UTF-8"><title>Vehicle Dashboard PDF</title><style>body{font-family:Arial,sans-serif;padding:24px;color:#172033}h1{margin:0 0 8px}table{width:100%;border-collapse:collapse;font-size:12px}th,td{border-bottom:1px solid #ddd;padding:8px;text-align:left}.summary{display:flex;gap:12px;margin:16px 0}.card{border:1px solid #ddd;border-radius:12px;padding:12px;min-width:160px}.value{font-size:24px;font-weight:800;color:#2563eb}</style></head><body><h1>Vehicle Dashboard</h1><div>Export: ${escapeHtml(printedAt)}</div><div class="summary"><div class="card">ยอดเก็บจริง<div class="value">${escapeHtml(totalAmount)}</div></div><div class="card">ยอดสุทธิตามระบบ<div class="value">${escapeHtml(netAmount)}</div></div></div><table><thead><tr><th>วันที่</th><th>ประเภทรถ</th><th>บริษัท</th><th>รายการ</th><th>ยอดสุทธิ</th><th>ยอดเก็บจริง</th></tr></thead><tbody>${rows.map(r=>`<tr><td>${escapeHtml(r.date)}</td><td>${escapeHtml(r.type)}</td><td>${escapeHtml(r.company)}</td><td>${escapeHtml(r.item)}</td><td>${escapeHtml(r.net_amount)}</td><td>${escapeHtml(r.collected_amount)}</td></tr>`).join('')}</tbody></table></body></html>`;
 const win=window.open('', '_blank');
 if(!win){alert('Browser บล็อก popup กรุณาอนุญาต popup แล้วลอง Export PDF อีกครั้ง');return}
 win.document.open();win.document.write(html);win.document.close();win.focus();setTimeout(()=>win.print(),700);
}
function showLoading(msg='กำลังโหลดข้อมูล...'){const t=box('loadingText'),o=box('loadingOverlay');if(t)t.textContent=msg;if(o)o.classList.add('show')}function hideLoading(){const o=box('loadingOverlay');if(o)o.classList.remove('show')}
async function load(){showLoading('กำลังค้นหาและอัปเดต Dashboard...');box('refreshStatus').textContent='กำลังโหลดข้อมูล...';try{const params=new URLSearchParams();const s=box('startDate').value,e=box('endDate').value,q=(box('searchBox')?box('searchBox').value.trim():'');if(s)params.set('start',s);if(e)params.set('end',e);if(q)params.set('q',q);const query=params.toString();const url='/api/dashboard'+(query?('?'+query+'&ts='+Date.now()):('?ts='+Date.now()));const res=await fetch(url).catch(()=>null);if(!res||!res.ok){box('status').textContent='ยังไม่มีข้อมูล';box('refreshStatus').textContent='ยังไม่มีข้อมูล';return}report=await res.json();allDays=report.dailyData;filteredDays=[...allDays];if(!s&&!e)setupRange();render(activeSelected);box('refreshStatus').textContent='ข้อมูลล่าสุดแล้ว • '+new Date().toLocaleTimeString('th-TH')+' • Auto refresh ทุก 15 นาที'}finally{hideLoading()}}
box('applyBtn').onclick=()=>load();box('resetBtn').onclick=()=>{setRange('','','');if(box('searchBox'))box('searchBox').value='';activeSelected='all';load()};box('showDateBtn').onclick=()=>{currentPage=1;renderCards(box('dateFilter').value)};box('showAllBtn').onclick=()=>{box('dateFilter').value='all';currentPage=1;renderCards('all')};box('dateFilter').onchange=()=>{currentPage=1;renderCards(box('dateFilter').value)};if(box('searchBox'))box('searchBox').oninput=()=>{currentPage=1;clearTimeout(window.searchTimer);window.searchTimer=setTimeout(()=>load(),450)};box('prevPageBtn').onclick=()=>{if(currentPage>1){currentPage--;renderCards(box('dateFilter').value)}};box('nextPageBtn').onclick=()=>{currentPage++;renderCards(box('dateFilter').value)};box('detailModeBtn').onclick=()=>{viewMode='detail';box('detailModeBtn').classList.add('active');box('compactModeBtn').classList.remove('active');renderCards(box('dateFilter').value)};box('compactModeBtn').onclick=()=>{viewMode='compact';box('compactModeBtn').classList.add('active');box('detailModeBtn').classList.remove('active');renderCards(box('dateFilter').value)};box('exportExcelBtn').onclick=exportExcel;box('exportPdfBtn').onclick=exportPDF;function initModernDatePicker(){if(window.flatpickr){rangePickerInstance=flatpickr('#rangePicker',{mode:'range',dateFormat:'Y-m-d',altInput:false,allowInput:false,disableMobile:true,onChange:(dates,txt,inst)=>{if(dates.length===2){setRange(fmtDate(dates[0]),fmtDate(dates[1]),'');load()}},onClose:(dates)=>{if(dates.length===1){setRange(fmtDate(dates[0]),fmtDate(dates[0]),'');load()}}});}}if(box('rangeOpenBtn'))box('rangeOpenBtn').onclick=()=>rangePickerInstance?.open();if(box('thisMonthBtn'))box('thisMonthBtn').onclick=()=>{setCurrentMonthDefault();load()};if(box('todayBtn'))box('todayBtn').onclick=()=>{setToday();load()};if(box('thisWeekBtn'))box('thisWeekBtn').onclick=()=>{setThisWeek();load()};initModernDatePicker();setCurrentMonthDefault();load();setInterval(()=>load(),900000);
</script></body></html>
"""



def read_html_file(filename: str, fallback: str) -> str:
    """Read an HTML file located beside app.py; fallback to embedded HTML if missing."""
    try:
        html_path = Path(__file__).resolve().parent / filename
        if html_path.exists():
            return html_path.read_text(encoding="utf-8")
    except Exception:
        pass
    return fallback


@app.get("/", include_in_schema=False)
def root() -> RedirectResponse:
    return RedirectResponse("/dashboard")


@app.get("/admin", response_class=HTMLResponse)
def admin_page() -> str:
    return read_html_file("admin.html", ADMIN_HTML)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard_page() -> str:
    return read_html_file("dashboard.html", DASHBOARD_HTML).replace("__APP_VERSION__", APP_VERSION)


@app.get("/executive", response_class=HTMLResponse)
def executive_page() -> str:
    return read_html_file("executive.html", "<h1>Executive Summary AI</h1><p>executive.html not found</p>")


@app.post("/api/import")
def api_import_legacy(raw_text: str = Form("")) -> JSONResponse:
    return api_import_text(raw_text=raw_text)


@app.post("/api/import/text")
def api_import_text(raw_text: str = Form("")) -> JSONResponse:
    text = raw_text.strip()
    if not text:
        raise HTTPException(status_code=400, detail="ไม่พบข้อมูลรายงาน")
    result = save_import_replace_all(text)
    return JSONResponse({"ok": True, "version": APP_VERSION, "import_type": "text", **result})


@app.post("/api/import/excel")
async def api_import_excel(file: UploadFile = File(...)) -> JSONResponse:
    name = file.filename or ""
    if not name.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="รองรับเฉพาะไฟล์ Excel .xlsx / .xls")
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="ไฟล์ Excel ว่าง")
    result = save_excel_replace_all(content, filename=name)
    return JSONResponse({"ok": True, **result})


@app.get("/api/dashboard")
def api_dashboard(
    start: str | None = Query(default=None),
    end: str | None = Query(default=None),
    q: str | None = Query(default=None),
) -> JSONResponse:
    cache_key = make_cache_key(start, end, q)
    cached = get_cached_dashboard(cache_key)
    if cached is not None:
        cached_copy = dict(cached)
        cached_copy["cache"] = {"hit": True, "ttl": CACHE_TTL_SECONDS}
        return JSONResponse(cached_copy)
    data = get_dashboard_data(start=start, end=end, q=q)
    data["cache"] = {"hit": False, "ttl": CACHE_TTL_SECONDS}
    set_cached_dashboard(cache_key, data)
    return JSONResponse(data)



@app.get("/api/health")
def api_health() -> JSONResponse:
    store, _ = read_github_store()
    return JSONResponse({
        "ok": True,
        "version": APP_VERSION,
        "storage": "github_json",
        "github_repo": GITHUB_REPO,
        "github_file": GITHUB_FILE,
        "github_branch": GITHUB_BRANCH,
        "daily_records": len(store.get("daily_records", [])),
        "weekly_summaries": len(store.get("weekly_summaries", [])),
        "updated_at": store.get("updated_at"),
        "cache_ttl_seconds": CACHE_TTL_SECONDS,
        "cache_key": DASHBOARD_CACHE.get("key"),
        "endpoints": {"excel": "/api/import/excel", "text": "/api/import/text", "legacy": "/api/import"},
    })


@app.post("/api/cache/clear")
def api_cache_clear(token: str = Form(...)) -> JSONResponse:
    if token != ADMIN_TOKEN:
        raise HTTPException(status_code=401, detail="Admin token ไม่ถูกต้อง")
    clear_dashboard_cache()
    return JSONResponse({"ok": True, "message": "cache cleared"})


@app.get("/api/github/test")
def api_github_test() -> JSONResponse:
    store, sha = read_github_store()
    return JSONResponse({
        "ok": True,
        "repo": GITHUB_REPO,
        "file": GITHUB_FILE,
        "branch": GITHUB_BRANCH,
        "sha_exists": bool(sha),
        "daily_records": len(store.get("daily_records", [])),
        "weekly_summaries": len(store.get("weekly_summaries", [])),
    })


@app.get("/api/report/latest")
def api_latest_report() -> JSONResponse:
    return api_dashboard()
